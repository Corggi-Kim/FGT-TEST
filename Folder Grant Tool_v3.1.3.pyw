# =====================================
# 1) 폴더 구조
#    project/
#      ├─ Folder_Grant_Tool.py
#      └─ assets/
#            ├─ logo.png
#            └─ fgt.ico
#
# 2) 실행 시 생성/사용 경로 (코드에서 자동 생성)
#    - C:\FGT\Log\Access : 실행 로그(access_YYYYMMDD.log)
#    - C:\FGT\conf      : 설정(login.json, theme 등)
#    - C:\FGT\ef         : 엑셀파일 다운로드
#    - C:\FGT\debug   : 디버그 파일
# =====================================

import sys, os, glob, datetime, subprocess, re, shutil, json, time
import html as htmllib
from typing import Dict, List
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import ElementClickInterceptedException, TimeoutException, StaleElementReferenceException
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QPushButton, QFileDialog, QTableWidget,
    QTableWidgetItem, QAbstractItemView, QVBoxLayout, QWidget, QTextEdit, QHBoxLayout,
    QLabel, QHeaderView, QCheckBox, QMenu, QToolButton, QMessageBox, QStyleOptionButton,
    QStyle, QDialog, QLineEdit, QComboBox, QDialogButtonBox, QFormLayout, QStatusBar,
    QSpinBox, QProgressBar, QAbstractButton
)
from PyQt5.QtGui import QFont, QGuiApplication, QKeySequence, QPainter, QTextOption, QTextCursor, QPixmap, QIcon
from PyQt5.QtCore import Qt, QRect, pyqtSignal, QThread, QObject, pyqtSlot, QProcess, QTimer, QPoint, QSize
from openpyxl import load_workbook

APP_NAME = "Folder Grant Tool"
APP_VERSION = "3.1.3"  #구조변경, 기능추가, 오류/버그수정
APP_BUILD = "2026-02-24"
APP_VERSION_STR = f"v{APP_VERSION}"

THEMES = {
    "light": {
        "bg": "#f5f5f5",
        "btn": "#d6dde0",
        "hover": "#b0bec5",
        "press": "#78909c",
        "panel": "#eceff1",
        "panel_border": "#cfd8dc",
        "select": "#d6e1e7",
        "alt": "#f6f9fb",
    },
    "dark":  {
        "bg": "#2f2f2f",
        "btn": "#424242",
        "hover": "#616161",
        "press": "#757575",
        "panel": "#3a3a3a",
        "panel_border": "#505050",
        "select": "#4a5a66",
        "alt": "#343838",
    },
}

HELP_TEXT = r"""[Folder Grant Tool 기능 설명]

1) 요청 확인
  - [🔍요청 확인]을 누르면 BUS의 폴더권한 요청 리스트를 자동으로 불러옵니다.
  - 리스트는 ‘구분(진행/종료)’, ‘요청(권한부여/권한해제)’ 으로 구분하여 진행과 종료과제를 모두 불러옵니다.
  - 리스트를 불러온 후 로그 창에 몇 건이 로드 되었는지 표시됩니다.
  - BUS 로그인 아이디/비밀번호가 필요한 경우 [⚙ 설정] 창이 먼저 뜹니다.

2) 설정
  - [⚙ 설정]에서 BUS 로그인 계정을 입력하고 저장 여부를 선택할 수 있습니다.
  - "저장"을 체크하면 다음 실행 시 계정을 다시 묻지 않습니다.

3) 파일 불러오기
  - 버스에서 다운로드 받은 엑셀파일을 임포트하여 수동으로 리스트를 불러올 수 있습니다.
  - [📂 파일 선택] 버튼을 클릭하여 불러오거나 엑셀 파일 드래그&드롭으로 로드합니다.

4) 수동 입력
  - [📝 수동 입력] 버튼으로 직접 데이터를 추가할 수 있습니다.
  - 사번, 프로젝트코드, Level2/Level3, Role 등을 입력합니다.

5) 행 선택 및 편집
  - 표 왼쪽 체크박스로 실행 대상을 고릅니다. 헤더의 체크박스로 전체 선택/해제 가능.
  - 더블클릭으로 일부 셀(구분/요청/사번/이름/프로젝트코드/Level2/Level3/부서/Role)을 수정할 수 있습니다.
  - 마우스 우클릭으로 [행 삭제], [선택된 행 삭제], [전체 삭제]를 할 수 있습니다.

6) 실행하기
  - [+ 실행] : 선택 행에 대해 권한 부여/해제 스크립트를 실행합니다.
     # 'Dry Run' 체크 시 실제 실행 없이 명령만 로그로 출력합니다.
     # '실행 후 자동 완료처리' 체크 시 폴더 권한부여 후 BUS 완료처리까지 자동 수행합니다.
  - 실행 중 [x 중지]로 전체 작업을 중단할 수 있습니다.
  - [v 완료 처리] : 간혹 실패한 BUS 완료처리를 수동으로 다시 처리할 수 있습니다.

7) 상태 값
  - 대기: 로드되어 폴더 권한을 부여하기 전
  - 검증필요/검증실패: 실행 전 점검에서 문제 발견
  - 실행중: 요청에 따른 폴더 권한 명령 실행 중
  - 추가완료/제거완료:  폴더 권한 명령 성공
  - 완료 처리중: BUS 완료처리 중
  - DryRun: 실제 실행 없이 명령만 출력
  - 실패/처리실패: 폴더 권한 명령 또는 BUS 처리 실패(로그/파일로 원인 확인)

8) 로그
  - 하단 로그 창에 요약과 결과가 표시됩니다(필터링된 핵심 메시지 위주).
  - 일자별 파일로 저장: C:\FGT\Log\Access\access_YYYYMMDD.log

9) 테마/로고/도움말
  - 우상단 🌙/🌞 버튼으로 라이트/다크 테마 전환.
  - 상단 로고/타이틀 클릭 시 이 도움말을 볼 수 있습니다.

10) 기타
  - Study/Isolated 및 STAT부서 Role에 따른 규칙을 자동 적용합니다.
  - 폴더 권한의 명령을 실행하는 과정에서 오래된 과제 폴더의 경우는 완료까지 오래 거릴 수 있습니다.
     특히, 종료과제 권한 해제의 경우 길게는 2시간 가량 소요됩니다. 사용 시 참고 바랍니다.
"""


LOG_DIR = r"C:\FGT\Log"
ACCESS_LOG_DIR = os.path.join(LOG_DIR, "Access")
REQUEST_TRACE_LOG_DIR = os.path.join(LOG_DIR, "RequestTrace")
CONF_DIR = r"C:\FGT\conf"
DL_DIR = r"C:\FGT\ef"
DEBUG_DIR = r"C:\FGT\debug"

GROUP_OU_PATH = r"OU=Group Project Folder,OU=1.Management Object Group,OU=lskglobal,DC=lskglobal,DC=com"
TEMPLATE_ROOT = r"\\LSK_S010\Study folder\_Template"

CONF_FILE = os.path.join(CONF_DIR, "login.json")
LOGO_FILE = "logo.png"
ICON_FILE = "fgt.ico"

os.makedirs(DL_DIR, exist_ok=True)
os.makedirs(LOG_DIR, exist_ok=True)
os.makedirs(ACCESS_LOG_DIR, exist_ok=True)
os.makedirs(REQUEST_TRACE_LOG_DIR, exist_ok=True)
os.makedirs(CONF_DIR, exist_ok=True)
os.makedirs(DEBUG_DIR, exist_ok=True)

def append_request_trace_log(tag: str, line: str):
    txt = (line or "") + "\n"
    fn = os.path.join(REQUEST_TRACE_LOG_DIR, f"{tag}_{datetime.date.today().strftime('%Y%m%d')}.log")
    try:
        os.makedirs(REQUEST_TRACE_LOG_DIR, exist_ok=True)
        with open(fn, "a", encoding="utf-8") as f:
            f.write(txt)
        return
    except Exception:
        pass
    try:
        os.makedirs(LOG_DIR, exist_ok=True)
        fallback = os.path.join(LOG_DIR, f"{tag}_{datetime.date.today().strftime('%Y%m%d')}.log")
        with open(fallback, "a", encoding="utf-8") as f:
            f.write(txt)
    except Exception:
        pass

DOMAIN_EMAIL_SUFFIX = "lskglobal.com"

BUS_LOGIN_URL = "https://bus.lskglobal.com/L4/Common/Login.aspx"
BUS_PROGRESS_LIST_URL = "https://bus.lskglobal.com/L4/Common/Default.aspx?7VHoVKC6bjriQDXQa/t/dQ=="
BUS_END_LIST_URL = "https://bus.lskglobal.com/L4/Common/Default.aspx?oRrCZc631pq4qaUZnht8Cg=="
BUS_NEW_URL = "https://bus.lskglobal.com/L4/Common/Default.aspx?2QiSDYhMfx5ql2mmNbos4A=="

REQ_GRANT = "권한부여"
REQ_RELEASE = "권한해제"
END_FLAG_GRANT = "END_GRANT"
END_FLAG_RELEASE = "END_RELEASE"

RELEASE_HINT_HEADERS = {
    "해제요청일", "해제요청자사번", "해제요청자성명",
    "해제요청자부서/팀", "해제요청자직책"
}

SHARE_ROOT = r"\\LSK_S010\Study Folder\{proj_seg}\{lv2}\{lv3}"
CLOSED_ROOT = r"\\192.168.1.95\Study_Closed"

LEVEL3_CHOICES = ["ARS","CO","DM","ER","MW","PM","PV","RA","SSU","STAT","STAT_IDMC","ETC"]

ROLE_MAP: Dict[str, List[str]] = {
    "Trial STAT/SP": ["3.Dataset", "4.Analysis", "5.SDTM", "6.Validation"],
    "Verification SP": ["3.Dataset", "6.Validation", "8.Verification"],
    "SDTM": ["3.Dataset", "5.SDTM", "6.Validation"],
    "Manager": ["3.Dataset", "4.Analysis", "5.SDTM", "6.Validation", "8.Verification"],
    "Randomization Statistician": ["Random"],
    "Blind Reviewer": ["Reviewer"],
    "Unblind Reviewer": ["Random", "Reviewer"],
}

STUDY_ROLES = {"Trial STAT/SP", "Verification SP", "SDTM", "Manager"}
ISOLATED_ROLES = {"Randomization Statistician", "Blind Reviewer", "Unblind Reviewer"}

ISOLATED_STAT_IDMC_ROLE_MAP: Dict[str, List[str]] = {
    "Trial STAT/SP": ["8.Verification"],
    "Verification SP": ["4.Analysis", "5.SDTM"],
    "SDTM": ["4.Analysis", "8.Verification"],
    "Manager": [],
}

LEGACY_STUDY_MAP = {
    "Trial STAT/SP": [3, 4, 5],
    "Verification SP": [3, 5, 8],
    "SDTM": [3, 5],
    "Manager": [3, 4, 5, 8],
}

NEW_THRESHOLD = 25069
FORCE_NEW_CODES = {
    "25-038", "25-032", "25-028", "25-006", "25-004",
    "24-084", "24-076", "24-072", "24-061", "24-037", "24-033", "24-026", "24-009",
    "21-040",
}
STAT_IDMC_NEW_THRESHOLD = 26012
STAT_IDMC_FORCE_NEW_CODES = {"25-074", "25-077"}

HEADER_ALIASES = {
    "user":   {"대상자사번"},
    "name": {"대상자성명"},
    "proj":   {"프로젝트코드"},
    "level1": {"폴더level1","폴더Level1"},
    "level2": {"폴더level2","폴더Level2"},
    "level3": {"폴더level3","폴더Level3"},
    "role":   {"statrole","STATROLE"},
    "dept":   {"대상자부서/팀"},
}

def build_root_from_proj(proj_raw: str) -> str:
    seg = proj_segment_for_folder(proj_raw)
    return r"\\LSK_S010\Study Folder\{}".format(seg)

def resource_path(rel_path: str) -> str:
    try:
        base = sys._MEIPASS
    except Exception:
        base = os.path.abspath(".")
    return os.path.join(base, rel_path)

def _norm_k(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "").replace("/", "").replace("_","")
RELEASE_HINT_HEADERS_NORM = {_norm_k(h) for h in RELEASE_HINT_HEADERS}

def _find_cols(header_row, wanted_norm_set):
    idxs = {}
    norm = [_norm_k(h) for h in header_row]
    for i, h in enumerate(norm):
        if h in wanted_norm_set and h not in idxs:
            idxs[h] = i
    return idxs

def _is_release_row_by_values(row_values, header_row):
    try:
        idxs = _find_cols(header_row, RELEASE_HINT_HEADERS_NORM)
        for k, i in idxs.items():
            if i < len(row_values):
                v = row_values[i]
                if v is not None and str(v).strip() != "":
                    return True
    except Exception:
        pass
    return False

def _norm_header(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "").replace("/", "").replace("_","").replace("-","")

def auto_map_columns(header_row):
    idx = {}
    norm_headers = [_norm_header(h if h is not None else "") for h in header_row]
    for key, aliases in HEADER_ALIASES.items():
        aliases_norm = {_norm_header(a) for a in aliases}
        for i, h in enumerate(norm_headers):
            if h in aliases_norm:
                idx[key] = i
                break
    return idx

def normalize_lv2(s: str) -> str:
    t = (s or "").strip().lower()
    return "Isolated" if "iso" in t else "Study"

def is_stat_idmc_lv3(lv3: str) -> bool:
    return (lv3 or "").strip().upper() == "STAT_IDMC"

def is_stat_lv3(lv3: str) -> bool:
    return (lv3 or "").strip().upper() in {"STAT", "STAT_IDMC"}

def is_lv3_etc(lv3: str) -> bool:
    return (lv3 or "").strip().lower() == "etc"

def insert_zero_middle_4digit(code4: str) -> str:
    return code4[:2] + "0" + code4[2:]

def split_proj_and_suffix(raw: str):
    s = (raw or "").strip()
    if "-" in s:
        base, suf = s.split("-", 1)
        return base.strip(), suf.strip()
    return s, ""

def proj_segment_for_folder(raw: str) -> str:
    base, suf = split_proj_and_suffix(raw)
    digits = re.sub(r"\D", "", base or "")
    if len(digits) == 4:
        digits = insert_zero_middle_4digit(digits)
    return digits + ("A" + suf if suf else "")

def group_digits_from_proj_for_groupname(raw: str) -> (str, str):
    base, suf = split_proj_and_suffix(raw)
    digits = re.sub(r"\D", "", base or "")
    if len(digits) == 4:
        digits = insert_zero_middle_4digit(digits)
    last5 = digits[-5:] if len(digits) >= 5 else digits.zfill(5)
    return last5, suf

def format_group_name(proj_raw: str, lv2: str) -> str:
    last5, suf = group_digits_from_proj_for_groupname(proj_raw)
    yy, xxx = last5[:2], last5[2:]
    name = f"LSK {yy}-{xxx}"
    if suf:
        name += f"-{suf}"
    if normalize_lv2(lv2) == "Isolated":
        name += " Isolated"
    return name

def _yyxxx_from_proj_for_groupname(proj_raw: str) -> str:
    last5, _ = group_digits_from_proj_for_groupname(proj_raw)
    return last5

def _lsk_code_for_compare(proj_raw: str) -> str:
    last5 = _yyxxx_from_proj_for_groupname(proj_raw)
    return f"{last5[:2]}-{last5[2:]}"

def is_new_template(proj_raw: str) -> bool:
    code_yyxxx = int(_yyxxx_from_proj_for_groupname(proj_raw))
    lsk_code = _lsk_code_for_compare(proj_raw)
    if lsk_code in FORCE_NEW_CODES:
        return True
    return code_yyxxx >= NEW_THRESHOLD

def is_stat_idmc_new_policy(proj_raw: str) -> bool:
    lsk_code = _lsk_code_for_compare(proj_raw)
    if lsk_code in STAT_IDMC_FORCE_NEW_CODES:
        return True
    try:
        code_yyxxx = int(_yyxxx_from_proj_for_groupname(proj_raw))
    except Exception:
        return False
    return code_yyxxx >= STAT_IDMC_NEW_THRESHOLD

def build_path_l3(proj_raw: str, lv2: str, lv3: str) -> str:
    seg = proj_segment_for_folder(proj_raw)
    return SHARE_ROOT.format(
        proj_seg=seg,
        lv2=normalize_lv2(lv2),
        lv3=(lv3 or "").strip().replace("\\","").replace("/","")
    )

def closed_segment_from_proj(proj_raw: str) -> str:
    s = (proj_raw or "").strip()
    m = re.fullmatch(r"(\d{5})(?:A(\d+))?$", s)
    if m:
        return s

    base, suf = split_proj_and_suffix(s)
    digits = re.sub(r"\D", "", base or "")
    if len(digits) == 4:
        digits = insert_zero_middle_4digit(digits)
    seg = digits[-5:].zfill(5)

    suf_digits = re.sub(r"\D", "", suf or "")
    return seg + (("A" + suf_digits) if suf_digits else "")

def build_closed_path_from_proj(proj_raw: str) -> str:
    return os.path.join(CLOSED_ROOT, closed_segment_from_proj(proj_raw))

def generate_add_script_closed(user_id: str, proj_raw: str) -> str:
    path = build_closed_path_from_proj(proj_raw)
    return f"icacls '{psq(path)}' /grant '{user_id}@{DOMAIN_EMAIL_SUFFIX}:(ci)(oi)rx';"

def generate_remove_script_closed(user_id: str, proj_raw: str) -> str:
    path = build_closed_path_from_proj(proj_raw)
    return f"icacls '{psq(path)}' /t /remove '{user_id}@{DOMAIN_EMAIL_SUFFIX}';"

def psq(s: str) -> str:
    return (s or "").replace("'", "''")

def pretty_cmd_lines(cmd: str) -> str:
    s = cmd
    s = re.sub(r';(?!\s*})', ';\n', s)
    s = re.sub(r'}\s*(?=if\b)', '}\n', s)
    s = re.sub(r'(?<!\n)\s*(?=\$fd\d+\s*=)', '\n', s)
    s = re.sub(r'(?<!\n)\s*(?=icacls\s+)', '\n', s, flags=re.IGNORECASE)
    s = re.sub(r'(?<!\n)\s*(?=(Add|Remove)-ADGroupMember\b)', '\n', s, flags=re.IGNORECASE)
    s = re.sub(r'[ \t]+\n', '\n', s)
    s = re.sub(r'(?m)^[ \t]+', '', s)
    s = re.sub(r'\n{2,}', '\n', s)
    return s.strip()

def _legacy_needed_nums(role: str):
    return LEGACY_STUDY_MAP.get(role, [])

def _legacy_find_missing_dirs(stat_path: str, needed_nums):
    missing = []
    try:
        for n in needed_nums:
            pattern = os.path.join(stat_path, f"{n}.*")
            candidates = [p for p in glob.glob(pattern) if os.path.isdir(p)]
            if not candidates:
                missing.append(n)
    except Exception:
        return needed_nums
    return missing

def build_legacy_study_add(user_id: str, stat_path: str, role: str) -> str:
    nums = _legacy_needed_nums(role)
    nums = [n for n in nums if n in (3, 4, 5, 8)]
    lines = []

    for n in nums:
        lines.append(
            f"$fd{n} = Get-ChildItem -Path '{psq(stat_path)}' -Directory | "
            f"Where-Object {{ $_.Name -like '{n}.*' }};"
        )
    lines.append(f"icacls '{psq(stat_path)}' /grant '{user_id}@{DOMAIN_EMAIL_SUFFIX}:(ci)(oi)rxm';")

    for n in nums:
        lines.append(
            f"if ($fd{n}) {{ icacls ('{psq(stat_path)}\\' + $fd{n}.Name) "
            f"/grant '{user_id}@{DOMAIN_EMAIL_SUFFIX}:(ci)(oi)rxm'; }}"
        )
    return " ".join(lines)

def build_legacy_study_remove(user_id: str, stat_path: str, role: str) -> str:
    nums = _legacy_needed_nums(role)
    nums = [n for n in nums if n in (3, 4, 5, 8)]
    lines = []
    for n in nums:
        lines.append(
            f"$fd{n} = Get-ChildItem -Path '{psq(stat_path)}' -Directory | "
            f"Where-Object {{ $_.Name -like '{n}.*' }};"
        )
    for n in nums:
        lines.append(
            f"if ($fd{n}) {{ icacls ('{psq(stat_path)}\\' + $fd{n}.Name) "
            f"/t /remove '{user_id}@{DOMAIN_EMAIL_SUFFIX}'; }}"
        )
    return " ".join(lines)

def build_legacy_isolated_add(user_id: str, path_l3: str) -> str:
    return f"icacls '{psq(path_l3)}' /grant '{user_id}@{DOMAIN_EMAIL_SUFFIX}:(ci)(oi)rxm';"

def build_legacy_isolated_remove(user_id: str, path_l3: str) -> str:
    return f"icacls '{psq(path_l3)}' /t /remove '{user_id}@{DOMAIN_EMAIL_SUFFIX}';"

def generate_add_script(user_id: str, proj_raw: str, lv2: str, lv3: str, role: str) -> str:
    group_name = format_group_name(proj_raw, lv2)
    path_l3 = build_path_l3(proj_raw, lv2, lv3)
    lv2_norm = normalize_lv2(lv2)
    role_clean = (role or "").strip()
    is_stat_idmc = is_stat_idmc_lv3(lv3)
    cmds = []

    cmds.append(
        f"try {{ Add-ADGroupMember -Identity '{psq(group_name)}' -Members '{psq(user_id)}' -ErrorAction Stop }} "
        f"catch {{ throw ('Add-ADGroupMember failed: ' + $_.Exception.Message) }};"
    )

    if is_lv3_etc(lv3):
        return ' '.join(cmds)

    is_new = is_new_template(proj_raw)

    if lv2_norm == "Study":
        if not role_clean:
            cmds.append(f"icacls '{psq(path_l3)}' /grant '{user_id}@{DOMAIN_EMAIL_SUFFIX}:(ci)(oi)rxm';")
            return ' '.join(cmds)
        if is_new:
            if role_clean not in STUDY_ROLES or role_clean not in ROLE_MAP:
                return ""
            cmds.append(f"icacls '{psq(path_l3)}' /grant '{user_id}@{DOMAIN_EMAIL_SUFFIX}:(ci)(oi)rxm';")
            for sub in ROLE_MAP[role_clean]:
                cmds.append(f"icacls '{psq(path_l3)}\\{psq(sub)}' /grant '{user_id}@{DOMAIN_EMAIL_SUFFIX}:(ci)(oi)rxm';")
        else:
            cmds.append(build_legacy_study_add(user_id, path_l3, role_clean))

    elif lv2_norm == "Isolated":
        if is_stat_idmc:
            if is_stat_idmc_new_policy(proj_raw):
                if not role_clean:
                    cmds.append(f"icacls '{psq(path_l3)}' /grant '{user_id}@{DOMAIN_EMAIL_SUFFIX}:(ci)(oi)rxm';")
                    return ' '.join(cmds)
                if role_clean and role_clean not in ISOLATED_STAT_IDMC_ROLE_MAP:
                    return ""
                if role_clean:
                    cmds.append(f"icacls '{psq(path_l3)}' /grant '{user_id}@{DOMAIN_EMAIL_SUFFIX}:(ci)(oi)rxm';")
                    blocked_subs = ISOLATED_STAT_IDMC_ROLE_MAP[role_clean]
                    blocked_expr = "@(" + ",".join([f"'{psq(s)}'" for s in blocked_subs]) + ")"
                    cmds.append(
                        f"Get-ChildItem -Path '{psq(path_l3)}' -Directory | "
                        f"Where-Object {{ {blocked_expr} -notcontains $_.Name }} | "
                        f"ForEach-Object {{ icacls $_.FullName /grant '{user_id}@{DOMAIN_EMAIL_SUFFIX}:(ci)(oi)rxm' }};"
                    )
            else:
                cmds.append(f"icacls '{psq(path_l3)}' /grant '{user_id}@{DOMAIN_EMAIL_SUFFIX}:(ci)(oi)rxm' /t;")
        elif is_new:
            if not role_clean:
                cmds.append(f"icacls '{psq(path_l3)}' /grant '{user_id}@{DOMAIN_EMAIL_SUFFIX}:(ci)(oi)rxm';")
                return ' '.join(cmds)
            if role_clean not in ISOLATED_ROLES:
                return ""
            iso_map = {
                "Randomization Statistician": ["Random"],
                "Blind Reviewer": ["Reviewer"],
                "Unblind Reviewer": ["Random", "Reviewer"],
            }
            cmds.append(f"icacls '{psq(path_l3)}' /grant '{user_id}@{DOMAIN_EMAIL_SUFFIX}:(ci)(oi)rxm';")
            for sub in iso_map[role_clean]:
                cmds.append(f"icacls '{psq(path_l3)}\\{psq(sub)}' /grant '{user_id}@{DOMAIN_EMAIL_SUFFIX}:(ci)(oi)rxm';")
        else:
            if not role_clean:
                cmds.append(f"icacls '{psq(path_l3)}' /grant '{user_id}@{DOMAIN_EMAIL_SUFFIX}:(ci)(oi)rxm';")
                return ' '.join(cmds)
            if role_clean != "Randomization Statistician":
                return ""
            cmds.append(build_legacy_isolated_add(user_id, path_l3))

    else:
        cmds.append(f"icacls '{psq(path_l3)}' /grant '{user_id}@{DOMAIN_EMAIL_SUFFIX}:(ci)(oi)rxm';")

    return ' '.join(cmds)

def generate_remove_script(user_id: str, proj_raw: str, lv2: str, lv3: str, role: str) -> str:
    group_name = format_group_name(proj_raw, lv2)
    path_l3 = build_path_l3(proj_raw, lv2, lv3)
    cmds = []
    cmds.append(f"Remove-ADGroupMember -Identity '{psq(group_name)}' -Members '{psq(user_id)}' -Confirm:$false;")

    if is_lv3_etc(lv3):
        return " ".join(cmds)

    cmds.append(f"icacls '{psq(path_l3)}' /t /remove '{user_id}@{DOMAIN_EMAIL_SUFFIX}';")
    return " ".join(cmds)

def _extract_tables_from_html(s: str):
    tables = re.findall(r"<table[^>]*>(.*?)</table>", s, re.I | re.S)
    out = []
    for tbl in tables:
        rows = re.findall(r"<tr[^>]*>(.*?)</tr>", tbl, re.I | re.S)
        parsed = []
        for r in rows:
            cells = re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", r, re.I | re.S)
            vals = []
            for c in cells:
                t = re.sub(r"<[^>]+>", "", c)
                t = htmllib.unescape(t).strip()
                vals.append(t)
            if vals:
                parsed.append(vals)
        if parsed:
            out.append(parsed)
    return out

def _parse_html_best_table(path: str):
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        html = f.read()
    candidates = _extract_tables_from_html(html)
    if not candidates:
        return [], []
    for tbl in candidates:
        header = tbl[0]
        data = tbl[1:]
        return header, data
    return [], []

class SettingsDialog(QDialog):
    def __init__(self, parent=None, saved=None, remembered=False, debug_on=False, fail_tol_default=5):
        try:
            self._fail_tol_default = int(fail_tol_default)
        except Exception:
            self._fail_tol_default = 5

        super().__init__(parent)
        self.setWindowTitle("로그인 계정")
        self.le_id = QLineEdit()
        self.le_pw = QLineEdit()
        self.le_pw.setEchoMode(QLineEdit.Password)
        if saved:
            self.le_id.setText(saved.get("id",""))
            self.le_pw.setText(saved.get("pw",""))

        self.chk_save = QCheckBox("저장")
        self.chk_save.setChecked(bool(remembered))
        
        self.chk_debug = QCheckBox("디버그 모드")
        self.chk_debug.setChecked(bool(debug_on))

        self.sb_tol = QSpinBox()
        self.sb_tol.setRange(0, 999)
        try:
            tol_saved = int((saved or {}).get("remove_fail_tol", self._fail_tol_default))
            self.sb_tol.setValue(tol_saved)
        except Exception:
            self.sb_tol.setValue(self._fail_tol_default)

        form = QFormLayout()

        self.sb_notify = QSpinBox()
        self.sb_notify.setRange(5, 60)
        self.sb_notify.setValue(int((saved or {}).get("notify_refresh_min", 10)))
        form.addRow("알림 리프레시(분)", self.sb_notify)

        form.addRow("아이디", self.le_id)
        form.addRow("비밀번호", self.le_pw)
        form.addRow("", self.chk_save)
        form.addRow("", self.chk_debug)
        form.addRow("폴더해제 실패 허용 개수", self.sb_tol)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)

        lay = QVBoxLayout(self)
        lay.addLayout(form)

        bottom_bar = QHBoxLayout()
        bottom_bar.setContentsMargins(0, 6, 0, 0)

        self.version_lbl = QLabel(f"{APP_VERSION_STR}")
        self.version_lbl.setObjectName("versionLabel")
        self.version_lbl.setStyleSheet("color:#8a8a8a; font-size:11px; padding:2px 4px;")
        self.version_lbl.setToolTip(f"Build: {APP_BUILD}")

        bottom_bar.addWidget(self.version_lbl, 0, Qt.AlignVCenter)
        bottom_bar.addStretch()
        bottom_bar.addWidget(btns)

        lay.addLayout(bottom_bar)

    def result(self):
        return (
            self.le_id.text().strip(),
            self.le_pw.text().strip(),
            self.chk_save.isChecked(),
            self.chk_debug.isChecked(),
            int(self.sb_tol.value()),
            int(self.sb_notify.value()),
        )

    def get_fail_tol(self) -> int:
        try:
            return int(self.sb_tol.value())
        except Exception:
            return 5

class BusSessionManager(QObject):
    readyChanged = pyqtSignal(bool, str)
    downloaded = pyqtSignal(str, str)
    busyChanged = pyqtSignal(bool)
    processed = pyqtSignal(object)
    newDownloaded = pyqtSignal(str, str)
    countsReady = pyqtSignal(dict)

    MAX_INIT_RETRY = 3
    INIT_RETRY_BASE_DELAY = 2.0     

    def __init__(self, dl_dir: str):
        super().__init__()
        self.dl_dir = dl_dir
        self.driver = None
        self._ready = False
        self._busy = False
        self._cancel = False
        self.debug_enabled = False
        self.debug_dir = DEBUG_DIR
        self._user = ""
        self._pw = ""

    def set_debug(self, enabled: bool, debug_dir: str | None = None):
        self.debug_enabled = bool(enabled)
        if debug_dir:
            self.debug_dir = debug_dir
        try:
            if self.debug_enabled:
                os.makedirs(self.debug_dir, exist_ok=True)
        except Exception:
            pass

    def _stabilize_grid(self, tries=2, nap=0.8):
        self._wait_overlay_gone(10)
        for _ in range(tries):
            self._click_search()
            self._wait_overlay_gone(10)
            self._go_iframe()
            time.sleep(nap)

    def set_creds(self, user: str, pw: str):
        self._user, self._pw = (user or "").strip(), (pw or "").strip()

    def is_ready(self) -> bool:
        return self._ready and self.driver is not None

    def is_busy(self) -> bool:
        return self._busy

    def _reset_filters(self):
        d = self.driver
        try:
            self._go_iframe()
            d.execute_script("""
              for (const id of ['approverYn','processYn','releaseYn']) {
                const el = document.getElementById(id);
                if (el) { el.value=''; el.dispatchEvent(new Event('change',{bubbles:true})); }
              }
            """)
        except Exception:
            pass

    def _nav_open_and_iframe(self, url: str, iframe_css: str = "iframe"):
        d = self.driver
        d.switch_to.default_content()
        d.get(url)
        WebDriverWait(d, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, iframe_css)))
        frame = d.find_element(By.CSS_SELECTOR, iframe_css)
        d.switch_to.frame(frame)

    def _set_select_value_and_fire(self, select_id: str, value: str):
        d = self.driver
        d.execute_script("""
            var sid = arguments[0], val = arguments[1];
            var s = document.getElementById(sid);
            if (s) {
                s.value = val;
                try { s.dispatchEvent(new Event('change', {bubbles:true})); } catch(e) {}
                if (typeof AllBtnYn === 'function') { try { AllBtnYn(); } catch(e) {} }
            }
        """, select_id, value)

    def _click_search_manual(self):
        d = self.driver
        try:
            WebDriverWait(d, 2).until(EC.element_to_be_clickable((By.ID, "btnSearch"))).click()
            return
        except Exception:
            pass

        d.switch_to.default_content()
        WebDriverWait(d, 5).until(EC.element_to_be_clickable((By.ID, "btnSearch"))).click()

        WebDriverWait(d, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "iframe")))
        d.switch_to.frame(d.find_element(By.CSS_SELECTOR, "iframe"))

    def _goto_new_site(self):
        try:
            d = self.driver
            d.switch_to.default_content()
            d.get(BUS_NEW_URL)
            WebDriverWait(d, 20).until(EC.presence_of_all_elements_located((By.TAG_NAME, "iframe")))
            self._go_iframe()
            self._wait_overlay_gone(10)
            time.sleep(0.8)
            return True
        except Exception:
            return False

    def _goto_progress_site(self):
        try:
            d = self.driver
            d.switch_to.default_content()
            d.get(BUS_PROGRESS_LIST_URL)
            WebDriverWait(d, 20).until(EC.presence_of_all_elements_located((By.TAG_NAME, "iframe")))
            self._go_iframe()
            self._wait_overlay_gone(10)
            time.sleep(1.0)
            return True
        except Exception:
            return False

    def _goto_end_site(self):
        try:
            d = self.driver
            d.switch_to.default_content()
            d.get(BUS_END_LIST_URL)
            WebDriverWait(d, 20).until(EC.presence_of_all_elements_located((By.TAG_NAME, "iframe")))
            self._go_iframe()
            self._wait_overlay_gone(10)
            time.sleep(1.0)
            return True
        except Exception:
            return False

    @pyqtSlot()
    def download_new_list(self):
        if not self.is_ready():
            self.newDownloaded.emit("", "세션 준비 안됨"); return
        if self._busy:
            self.newDownloaded.emit("", "다른 작업 실행중"); return

        self._busy = True
        self.busyChanged.emit(True)
        self._cancel = False
        d = self.driver

        try:
            self._goto_new_site()
            d = self.driver
            self._go_iframe()
            try:
                d.execute_script("""
                  const p=document.getElementById('processYn');
                  if(p){ p.value='N'; p.dispatchEvent(new Event('change',{bubbles:true})); }
                """)
                time.sleep(0.2)
                self._go_iframe()
                WebDriverWait(d, 5).until(EC.element_to_be_clickable((By.ID, "btnSearch"))).click()
            except Exception:
                pass

            before = set(glob.glob(os.path.join(self.dl_dir, "*.xls"))) | set(glob.glob(os.path.join(self.dl_dir, "*.xlsx")))

            WebDriverWait(d, 10).until(EC.element_to_be_clickable((By.ID, "btnExcel"))).click()
            try:
                WebDriverWait(d, 8).until(EC.visibility_of_element_located((By.ID, "txtExcelDownReason")))
                try:
                    el = d.find_element(By.ID, "txtExcelDownReason")
                    el.clear(); el.send_keys("신규확인-미완료(N) 목록")
                except Exception:
                    pass
                ok_btn = WebDriverWait(d, 6).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button.swal2-confirm")))
                try: ok_btn.click()
                except ElementClickInterceptedException:
                    d.execute_script("arguments[0].click();", ok_btn)
            except TimeoutException:
                pass

            end = time.time() + 120
            latest = ""
            while time.time() < end:
                self._check_cancel()
                if not list(glob.glob(os.path.join(self.dl_dir, "*.crdownload"))):
                    now = set(glob.glob(os.path.join(self.dl_dir, "*.xls"))) | set(glob.glob(os.path.join(self.dl_dir, "*.xlsx")))
                    new_files = list(now - before)
                    if new_files:
                        latest = max(new_files, key=os.path.getmtime)
                        break
                time.sleep(0.3)

            if not latest:
                self.newDownloaded.emit("", "다운로드 실패"); return

            out_path = os.path.join(self.dl_dir, "신규_미완료.xls")
            try:
                if os.path.exists(out_path): os.remove(out_path)
            except Exception:
                pass
            try:
                shutil.move(latest, out_path)
            except Exception:
                try:
                    shutil.copyfile(latest, out_path)
                    try: os.remove(latest)
                    except Exception: pass
                except Exception:
                    self.newDownloaded.emit("", "파일 저장 실패"); return

            self.newDownloaded.emit(out_path, "")
        except SystemExit:
            self.newDownloaded.emit("", "사용자 취소")
        except Exception as e:
            self.newDownloaded.emit("", f"오류: {e}")
        finally:
            self._busy = False
            self.busyChanged.emit(False)

    def _read_total_entries(self) -> int:
        d = self.driver
        try:
            info_el = WebDriverWait(d, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".dataTables_info"))
            )
            txt = (info_el.text or "").strip().lower()
            m = re.search(r"(\d+)", txt)
            if m:
                return int(m.group(1))
        except Exception:
            pass

        try:
            self._stabilize_grid()
            WebDriverWait(d, 5).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "tbody tr")))
            return len(d.find_elements(By.CSS_SELECTOR, "tbody tr"))
        except Exception:
            return 0

    def collect_request_counts(self):
        if not self.is_ready():
            try:
                self.countsReady.emit(getattr(self, "_last_counts", {}))
            finally:
                return
        if getattr(self, "_busy", False):

            try:
                self.countsReady.emit(getattr(self, "_last_counts", {}))
            finally:
                return

        d = self.driver
        counts = {
            "신규미완료": 0,
            "진행-부여": 0,
            "진행-제거": 0,
            "종료-부여": 0,
            "종료-제거": 0,
        }

        try:
            try:
                self._nav_open_and_iframe(BUS_NEW_URL)

                try:
                    pre = d.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip()
                except Exception:
                    pre = ""

                self._set_select_value_and_fire("processYn", "N")
                self._click_search_manual()

                try:
                    WebDriverWait(d, 7).until(
                        lambda x: (x.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip()
                                   if x.find_elements(By.CSS_SELECTOR, ".dataTables_info") else "") != (pre or "")
                    )
                except Exception:
                    time.sleep(0.6)

                txt = d.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip().lower()
                m = re.search(r"(\d+)", txt)
                counts["신규미완료"] = int(m.group(1)) if m else 0
            except Exception:
                pass

            try:
                self._nav_open_and_iframe(BUS_PROGRESS_LIST_URL)

                try:
                    pre = d.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip()
                except Exception:
                    pre = ""

                self._set_select_value_and_fire("processYn", "N")
                self._click_search_manual()

                try:
                    WebDriverWait(d, 7).until(
                        lambda x: (x.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip()
                                   if x.find_elements(By.CSS_SELECTOR, ".dataTables_info") else "") != (pre or "")
                    )
                except Exception:
                    time.sleep(0.6)

                txt = d.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip().lower()
                m = re.search(r"(\d+)", txt)
                counts["진행-부여"] = int(m.group(1)) if m else 0
            except Exception:
                pass

            try:
                self._nav_open_and_iframe(BUS_PROGRESS_LIST_URL)

                try:
                    pre = d.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip()
                except Exception:
                    pre = ""

                self._set_select_value_and_fire("releaseYn", "N")
                self._click_search_manual()

                try:
                    WebDriverWait(d, 7).until(
                        lambda x: (x.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip()
                                   if x.find_elements(By.CSS_SELECTOR, ".dataTables_info") else "") != (pre or "")
                    )
                except Exception:
                    time.sleep(0.6)

                txt = d.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip().lower()
                m = re.search(r"(\d+)", txt)
                counts["진행-제거"] = int(m.group(1)) if m else 0
            except Exception:
                pass

            try:
                self._nav_open_and_iframe(BUS_END_LIST_URL)

                try:
                    pre = d.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip()
                except Exception:
                    pre = ""

                self._set_select_value_and_fire("processYn", "N")
                self._click_search_manual()

                try:
                    WebDriverWait(d, 7).until(
                        lambda x: (x.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip()
                                   if x.find_elements(By.CSS_SELECTOR, ".dataTables_info") else "") != (pre or "")
                    )
                except Exception:
                    time.sleep(0.6)

                txt = d.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip().lower()
                m = re.search(r"(\d+)", txt)
                counts["종료-부여"] = int(m.group(1)) if m else 0
            except Exception:
                pass

            try:
                self._nav_open_and_iframe(BUS_END_LIST_URL)

                try:
                    pre = d.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip()
                except Exception:
                    pre = ""

                self._set_select_value_and_fire("releaseYn", "N")
                self._click_search_manual()

                try:
                    WebDriverWait(d, 7).until(
                        lambda x: (x.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip()
                                   if x.find_elements(By.CSS_SELECTOR, ".dataTables_info") else "") != (pre or "")
                    )
                except Exception:
                    time.sleep(0.6)

                txt = d.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip().lower()
                m = re.search(r"(\d+)", txt)
                counts["종료-제거"] = int(m.group(1)) if m else 0
            except Exception:
                pass

        finally:
            self._last_counts = counts
            try:
                self.countsReady.emit(counts)
            except Exception:
                pass
            
    @pyqtSlot()
    def start(self):
        def _cleanup_driver(drv):
            if not drv:
                return
            try:
                drv.quit()
            except Exception:
                try:
                    if hasattr(drv, "service") and drv.service and drv.service.process:
                        drv.service.process.kill()
                except Exception:
                    pass

        if not (self._user and self._pw):
            self._ready = False
            self.readyChanged.emit(False, "자격증명 없음")
            return

        if self.driver:
            self._ready = True
            self.readyChanged.emit(True, "이미 준비됨")
            return

        attempt = 0
        last_err = None

        while attempt < self.MAX_INIT_RETRY and not self._cancel:
            attempt += 1
            try:
                options = webdriver.ChromeOptions()
                prefs = {
                    "download.default_directory": self.dl_dir,
                    "download.prompt_for_download": False,
                    "download.directory_upgrade": True,
                    "safebrowsing.enabled": True,
                }
                options.add_experimental_option("prefs", prefs)
                options.add_argument("--headless=new")
                options.add_argument("--disable-gpu")
                options.add_argument("--no-sandbox")
                options.add_argument("--window-size=1280,900")

                self.driver = webdriver.Chrome(options=options)
                d = self.driver

                d.get(BUS_LOGIN_URL)
                WebDriverWait(d, 20).until(EC.presence_of_element_located((By.ID, "windowsaccount")))
                d.find_element(By.ID, "windowsaccount").send_keys(self._user)
                d.find_element(By.ID, "password").send_keys(self._pw)
                d.find_element(By.ID, "btnLogin").click()
                WebDriverWait(d, 20).until(EC.url_contains("Common"))

                try:
                    d.switch_to.window(d.window_handles[-1])
                except Exception:
                    pass

                d.get(BUS_PROGRESS_LIST_URL)
                WebDriverWait(d, 20).until(EC.presence_of_all_elements_located((By.TAG_NAME, "iframe")))
                self._go_iframe()

                WebDriverWait(d, 10).until(EC.element_to_be_clickable((By.ID, "approverYn")))
                Select(d.find_element(By.ID, "approverYn")).select_by_value("1")
                
                self._set_request_filter(REQ_GRANT)

                self._ready = True
                self.readyChanged.emit(True, f"준비됨(시도 {attempt}/{self.MAX_INIT_RETRY})")
                return

            except Exception as e:
                last_err = e
                _cleanup_driver(self.driver)
                self.driver = None
                self._ready = False
                self.readyChanged.emit(False, f"세션 초기화 실패({attempt}/{self.MAX_INIT_RETRY}): {e}")

                if attempt >= self.MAX_INIT_RETRY or self._cancel:
                    break

                delay = self.INIT_RETRY_BASE_DELAY * (2 ** (attempt - 1))
                end = time.time() + delay
                while time.time() < end and not self._cancel:
                    QApplication.processEvents()
                    time.sleep(0.05)

        self._ready = False
        self.readyChanged.emit(False, f"세션 초기화 최종 실패: {last_err}")

    @pyqtSlot()
    def stop(self):
        if getattr(self, "_stopping", False):
            return
        self._stopping = True
        try:
            self._cancel = True
            self._busy = False
            self._ready = False
            self.busyChanged.emit(False)
            self.readyChanged.emit(False, "정지됨")

            d = self.driver
            self.driver = None
            if d:
                try:
                    d.quit()
                except Exception:
                    try:
                        if hasattr(d, "service") and d.service and d.service.process:
                            d.service.process.kill()
                    except Exception:
                        pass
        finally:
            self._stopping = False

    @pyqtSlot()
    def cancel_current(self):
        self._cancel = True

    def _check_cancel(self, context: str = "download"):
        if self._cancel:
            self._busy = False
            self.busyChanged.emit(False)
            self._cancel = False
            if context == "process":
                self.processed.emit([])
            else:
                self.downloaded.emit("", "사용자 취소")
            raise SystemExit

    @pyqtSlot()
    def download_list(self):
        if not self.is_ready():
            self.downloaded.emit("", "세션 준비 안됨")
            return
        if self._busy:
            self.downloaded.emit("", "다른 작업 실행중")
            return

        self._busy = True
        self.busyChanged.emit(True)
        self._cancel = False

        d = self.driver

        def _search_and_download_end(req_flag: str) -> str:
            self._goto_end_site()
            self._set_end_filter(req_flag)

            d = self.driver
            try:
                self._go_iframe()
                try:
                    WebDriverWait(d, 5).until(EC.element_to_be_clickable((By.ID, "btnSearch"))).click()
                except Exception:
                    d.switch_to.default_content()
                    WebDriverWait(d, 5).until(EC.element_to_be_clickable((By.ID, "btnSearch"))).click()
                    self._go_iframe()
            except Exception:
                pass

            before = set(glob.glob(os.path.join(self.dl_dir, "*.xls"))) | set(glob.glob(os.path.join(self.dl_dir, "*.xlsx")))

            try:
                self._go_iframe()
                WebDriverWait(d, 10).until(EC.element_to_be_clickable((By.ID, "btnExcel"))).click()
            except Exception:
                d.switch_to.default_content()
                WebDriverWait(d, 10).until(EC.element_to_be_clickable((By.ID, "btnExcel"))).click()

            try:
                WebDriverWait(d, 10).until(EC.visibility_of_element_located((By.ID, "txtExcelDownReason")))
                try:
                    reason = "종료권한리스트 수집-" + ("해제" if req_flag == END_FLAG_RELEASE else "부여")
                    el = d.find_element(By.ID, "txtExcelDownReason")
                    el.clear(); el.send_keys(reason)
                except Exception:
                    pass
                ok_btn = WebDriverWait(d, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button.swal2-confirm")))
                try: ok_btn.click()
                except ElementClickInterceptedException:
                    d.execute_script("arguments[0].click();", ok_btn)
            except TimeoutException:
                pass

            end = time.time() + 120
            latest = ""
            while time.time() < end:
                self._check_cancel()
                if not list(glob.glob(os.path.join(self.dl_dir, "*.crdownload"))):
                    now = set(glob.glob(os.path.join(self.dl_dir, "*.xls"))) | set(glob.glob(os.path.join(self.dl_dir, "*.xlsx")))
                    new_files = list(now - before)
                    if new_files:
                        latest = max(new_files, key=os.path.getmtime)
                        break
                time.sleep(0.35)

            if not latest:
                return ""

            suffix = "해제" if req_flag == END_FLAG_RELEASE else "부여"
            out_path = os.path.join(self.dl_dir, f"종료권한리스트_{suffix}.xls")
            try:
                if os.path.exists(out_path): os.remove(out_path)
            except Exception:
                pass

            try:
                shutil.move(latest, out_path)
            except Exception:
                try:
                    shutil.copyfile(latest, out_path)
                    try: os.remove(latest)
                    except Exception: pass
                except Exception:
                    return ""

            return out_path

        def _search_and_download(reqtype: str) -> str:
            self._goto_progress_site()
            self._set_request_filter(reqtype)
            try:
                self._go_iframe()
                try:
                    WebDriverWait(d, 5).until(EC.element_to_be_clickable((By.ID, "btnSearch"))).click()
                except Exception:
                    d.switch_to.default_content()
                    WebDriverWait(d, 5).until(EC.element_to_be_clickable((By.ID, "btnSearch"))).click()
                    self._go_iframe()
            except Exception:
                pass

            before = set(glob.glob(os.path.join(self.dl_dir, "*.xls"))) | set(
                glob.glob(os.path.join(self.dl_dir, "*.xlsx"))
            )

            WebDriverWait(d, 10).until(EC.element_to_be_clickable((By.ID, "btnExcel"))).click()
            WebDriverWait(d, 10).until(EC.visibility_of_element_located((By.ID, "txtExcelDownReason")))
            try:
                reason = f"권한리스트 수집-{'해제' if reqtype == REQ_RELEASE else '부여'}"
                el = d.find_element(By.ID, "txtExcelDownReason")
                el.clear()
                el.send_keys(reason)
            except Exception:
                pass

            ok_btn = WebDriverWait(d, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "button.swal2-confirm"))
            )
            try:
                ok_btn.click()
            except ElementClickInterceptedException:
                d.execute_script("arguments[0].click();", ok_btn)

            end = time.time() + 120
            latest = ""
            while time.time() < end:
                self._check_cancel()
                if not list(glob.glob(os.path.join(self.dl_dir, "*.crdownload"))):
                    now = set(glob.glob(os.path.join(self.dl_dir, "*.xls"))) | set(
                        glob.glob(os.path.join(self.dl_dir, "*.xlsx"))
                    )
                    new_files = list(now - before)
                    if new_files:
                        latest = max(new_files, key=os.path.getmtime)
                        break
                time.sleep(0.35)

            if not latest:
                return ""

            suffix = "해제" if reqtype == REQ_RELEASE else "부여"
            out_path = os.path.join(self.dl_dir, f"권한리스트_{suffix}.xls")
            try:
                if os.path.exists(out_path):
                    os.remove(out_path)
            except Exception:
                pass

            try:
                shutil.move(latest, out_path)
            except Exception:
                try:
                    shutil.copyfile(latest, out_path)
                    try:
                        os.remove(latest)
                    except Exception:
                        pass
                except Exception:
                    return ""

            return out_path

        try:
            for f in glob.glob(os.path.join(self.dl_dir, "*")):
                try:
                    os.remove(f)
                except:
                    pass

            grant_path = _search_and_download(REQ_GRANT)
            release_path = _search_and_download(REQ_RELEASE)

            end_grant_path   = _search_and_download_end(END_FLAG_GRANT)
            end_release_path = _search_and_download_end(END_FLAG_RELEASE)

            header_g, data_g = _parse_html_best_table(grant_path) if grant_path else ([], [])
            header_r, data_r = _parse_html_best_table(release_path) if release_path else ([], [])
            has_normal = bool(grant_path or release_path)
            combined_path = ""
            if has_normal:
                header = header_g if len(header_g) >= len(header_r) else header_r
                def _pad(row, n): rr = list(row);  rr += [""] * max(0, n-len(rr)); return rr[:n]
                rows = []
                for r in data_g: rows.append(_pad(r, len(header)))
                for r in data_r: rows.append(_pad(r, len(header)))
                parts = ["<html><head><meta charset='utf-8'></head><body>",
                         "<table border='1'>",
                         "<thead><tr>" + "".join(f"<th>{htmllib.escape(str(h or ''))}</th>" for h in header) + "</tr></thead>",
                         "<tbody>"]
                for r in rows:
                    parts.append("<tr>" + "".join(f"<td>{htmllib.escape(str(c or ''))}</td>" for c in r) + "</tr>")
                parts.append("</tbody></table></body></html>")
                combined_path = os.path.join(self.dl_dir, "권한리스트_합본.xls")
                with open(combined_path, "w", encoding="utf-8") as f:
                    f.write("\n".join(parts))

            header_eg, data_eg = _parse_html_best_table(end_grant_path) if end_grant_path else ([], [])
            header_er, data_er = _parse_html_best_table(end_release_path) if end_release_path else ([], [])
            has_end = bool(end_grant_path or end_release_path)
            combined_end_path = ""
            if has_end:
                header_e = header_eg if len(header_eg) >= len(header_er) else header_er
                def _pad_e(row, n):
                    rr = list(row); rr += [""] * max(0, n-len(rr)); return rr[:n]
                rows_e = []
                for r in data_eg: rows_e.append(_pad_e(r, len(header_e)))
                for r in data_er: rows_e.append(_pad_e(r, len(header_e)))
                parts_e = ["<html><head><meta charset='utf-8'></head><body>",
                           "<table border='1'>",
                           "<thead><tr>" + "".join(f"<th>{htmllib.escape(str(h or ''))}</th>" for h in header_e) + "</tr></thead>",
                           "<tbody>"]
                for r in rows_e:
                    parts_e.append("<tr>" + "".join(f"<td>{htmllib.escape(str(c or ''))}</td>" for c in r) + "</tr>")
                parts_e.append("</tbody></table></body></html>")
                combined_end_path = os.path.join(self.dl_dir, "종료권한리스트_합본.xls")
                with open(combined_end_path, "w", encoding="utf-8") as f:
                    f.write("\n".join(parts_e))

            if not has_normal and not has_end:
                self.downloaded.emit("", "부여/해제 다운로드 실패")
                return

            primary = combined_path if has_normal else combined_end_path
            self.downloaded.emit(primary, "")

        except SystemExit:
            self.downloaded.emit("", "사용자 취소")
        except Exception as e:
            self.downloaded.emit("", f"오류: {e}")
        finally:
            self._busy = False
            self.busyChanged.emit(False)

    def _ensure_iframe(self):
        try:
            d = self.driver
            d.switch_to.default_content()
            frames = d.find_elements(By.TAG_NAME, "iframe")
            if frames:
                d.switch_to.frame(frames[0])
            return True
        except Exception:
            return False

    def _go_iframe(self):
        try:
            d = self.driver
            d.switch_to.default_content()
            frames = d.find_elements(By.TAG_NAME, "iframe")
            if frames:
                d.switch_to.frame(frames[0])
            return True
        except Exception:
            return False

    def _wait_overlay_gone(self, timeout=10):
        d = self.driver
        try:
            WebDriverWait(d, timeout).until(
                EC.invisibility_of_element_located(
                    (By.CSS_SELECTOR, ".swal2-container, .loading, .blockUI, .blockOverlay")
                )
            )
        except Exception:
            pass
            
    def _debug_dump(self, tag: str):
        if not self.debug_enabled:
            return
        try:
            d = self.driver
            ts = int(time.time())
            base = f"debug_{tag}_{ts}"
            html_path = os.path.join(self.debug_dir, f"{base}.html")
            png_path  = os.path.join(self.debug_dir, f"{base}.png")
            with open(html_path, "w", encoding="utf-8") as f:
                f.write(d.page_source)
            try:
                d.save_screenshot(png_path)
            except Exception:
                pass
        except Exception:
            pass
        
    def _set_request_filter(self, reqtype: str):
        self._goto_progress_site()
        d = self.driver
        self._go_iframe()

        try:
            d.execute_script("""
              for (const id of ['approverYn','processYn','releaseYn']) {
                const el = document.getElementById(id);
                if (el) { el.value=''; el.dispatchEvent(new Event('change',{bubbles:true})); }
              }
            """)
        except Exception:
            pass

        is_release = (reqtype or "").strip() in (REQ_RELEASE, "권한해제", "해제", "release", "rel")

        try:
            if is_release:
                d.execute_script("""
                  const a=document.getElementById('approverYn');
                  const p=document.getElementById('processYn');
                  const r=document.getElementById('releaseYn');
                  if (a) { a.value='';  a.dispatchEvent(new Event('change',{bubbles:true})); }
                  if (p) { p.value='';  p.dispatchEvent(new Event('change',{bubbles:true})); }
                  if (r) { r.value='N'; r.dispatchEvent(new Event('change',{bubbles:true})); }
                """)
            else:
                d.execute_script("""
                  const a=document.getElementById('approverYn');
                  const p=document.getElementById('processYn');
                  const r=document.getElementById('releaseYn');
                  if (a) { a.value='1'; a.dispatchEvent(new Event('change',{bubbles:true})); }
                  if (p) { p.value='N'; p.dispatchEvent(new Event('change',{bubbles:true})); }
                  if (r) { r.value='';  r.dispatchEvent(new Event('change',{bubbles:true})); }
                """)
        except Exception:
            pass

        time.sleep(0.3)
        try:
            vals = d.execute_script("""
              const g = id => (document.getElementById(id)||{}).value || '';
              return {a:g('approverYn'), p:g('processYn'), r:g('releaseYn')};
            """)
            if is_release:
                ok = (vals.get('a','') == '' and vals.get('p','') == '' and vals.get('r','') == 'N')
            else:
                ok = (vals.get('a','') == '1' and vals.get('p','') == 'N' and vals.get('r','') == '')
            if not ok:
                return self._set_request_filter(reqtype)
        except Exception:
            pass

        self._debug_dump("progress_release_filter" if is_release else "progress_grant_filter")

    def _set_end_filter(self, reqtype_flag: str):
        self._goto_end_site()
        d = self.driver
        self._go_iframe()

        try:
            d.execute_script("""
              for (const id of ['approverYn','processYn','releaseYn']) {
                const el = document.getElementById(id);
                if (el) { el.value=''; el.dispatchEvent(new Event('change',{bubbles:true})); }
              }
            """)
        except Exception:
            pass

        try:
            if reqtype_flag == END_FLAG_GRANT:
                d.execute_script("""
                  const a=document.getElementById('approverYn');
                  const p=document.getElementById('processYn');
                  const r=document.getElementById('releaseYn');
                  if (a) { a.value='1'; a.dispatchEvent(new Event('change',{bubbles:true})); }
                  if (p) { p.value='N'; p.dispatchEvent(new Event('change',{bubbles:true})); }
                  if (r) { r.value='';  r.dispatchEvent(new Event('change',{bubbles:true})); }
                """)
            else:
                d.execute_script("""
                  const a=document.getElementById('approverYn');
                  const p=document.getElementById('processYn');
                  const r=document.getElementById('releaseYn');
                  if (a) { a.value='';  a.dispatchEvent(new Event('change',{bubbles:true})); }
                  if (p) { p.value='';  p.dispatchEvent(new Event('change',{bubbles:true})); }
                  if (r) { r.value='N'; r.dispatchEvent(new Event('change',{bubbles:true})); }
                """)
        except Exception:
            pass

        time.sleep(0.3)
        try:
            vals = d.execute_script("""
              const g = id => (document.getElementById(id)||{}).value || '';
              return {a:g('approverYn'), p:g('processYn'), r:g('releaseYn')};
            """)
            if reqtype_flag == END_FLAG_GRANT:
                ok = (vals.get('a','') == '1' and vals.get('p','') == 'N' and vals.get('r','') == '')
            else:
                ok = (vals.get('a','') == '' and vals.get('p','') == '' and vals.get('r','') == 'N')
            if not ok:
                return self._set_end_filter(reqtype_flag)
        except Exception:
            pass

        self._debug_dump("end_grant_filter" if reqtype_flag == END_FLAG_GRANT else "end_release_filter")

    def _click_search(self):
        d = self.driver
        try:
            self._go_iframe()
            try:
                WebDriverWait(d, 4).until(EC.element_to_be_clickable((By.ID, "btnSearch"))).click()
            except Exception:
                d.switch_to.default_content()
                WebDriverWait(d, 4).until(EC.element_to_be_clickable((By.ID, "btnSearch"))).click()
            self._wait_overlay_gone(10)
            self._go_iframe()
        except Exception:
            pass

    def _for_each_page(self, work_on_page):
        d = self.driver
        tried = 0

        self._check_cancel("process")
        if work_on_page():
            return True

        while tried < 200:
            self._check_cancel("process")            
            self._go_iframe()
            next_btn = None
            try:
                cands = d.find_elements(By.CSS_SELECTOR, "a[aria-label='Next'], button[aria-label='Next']")
                if not cands:
                    cands = d.find_elements(By.XPATH, "//a[normalize-space()='다음' or normalize-space()='>'] | //button[normalize-space()='다음' or normalize-space()='>']")
                if cands:
                    next_btn = cands[-1]
                    cls = (next_btn.get_attribute("class") or "").lower()
                    aria = (next_btn.get_attribute("aria-disabled") or "").lower()
                    if "disabled" in cls or aria in ("true", "1"):
                        break
            except Exception:
                next_btn = None

            if not next_btn:
                break

            try:
                d.execute_script("arguments[0].scrollIntoView({block:'center'});", next_btn)
            except Exception:
                pass

            self._check_cancel("process")
            try:
                next_btn.click()
            except ElementClickInterceptedException:
                d.execute_script("arguments[0].click();", next_btn)

            self._wait_overlay_gone(10)
            self._go_iframe()
            time.sleep(0.4)

            self._check_cancel("process")
            if work_on_page():
                return True

            tried += 1

        return False

    def _norm(self, s):
        return (s or "").strip().replace("\u00a0", " ").lower()

    def _bus_code_variants(self, proj_raw: str):
        try:
            code5 = _yyxxx_from_proj_for_groupname(proj_raw)
            c5_int = int(code5)
            code4 = None
            if c5_int < 20000 and len(code5) == 5 and code5[2] == "0":
                code4 = code5[:2] + code5[3:]
            return code5, code4
        except Exception:
            base = re.sub(r"\D", "", str(proj_raw or ""))
            if len(base) >= 5:
                base = base[-5:]
            return base, None

    def _goto_first_page(self):
        d = self.driver
        try:
            self._go_iframe()
            btns = d.find_elements(By.XPATH, "//a[normalize-space()='처음'] | //button[normalize-space()='처음']")
            if btns:
                try:
                    d.execute_script("arguments[0].scrollIntoView({block:'center'});", btns[0])
                except Exception:
                    pass
                try:
                    btns[0].click()
                except ElementClickInterceptedException:
                    d.execute_script("arguments[0].click();", btns[0])
                self._wait_overlay_gone(10)
                self._go_iframe()
                return
            
            page1 = d.find_elements(By.XPATH, "//a[normalize-space()='1'] | //button[normalize-space()='1']")
            if page1:
                try:
                    d.execute_script("arguments[0].scrollIntoView({block:'center'});", page1[0])
                except Exception:
                    pass
                try:
                    page1[0].click()
                except ElementClickInterceptedException:
                    d.execute_script("arguments[0].click();", page1[0])
                self._wait_overlay_gone(10)
                self._go_iframe()
        except Exception:
            pass

    def _narrow_by_path_strict(self, rows, lv2, lv3, path_hint):
        if not rows:
            return rows

        def N(s): return (s or "").replace("\u00a0"," ").strip().lower()
        lv2n = N(lv2)
        lv3n = N(lv3)
        tail = (path_hint or "").split("\\")[-1]
        tailn = N(tail)

        must_tokens = [t for t in [lv2n, lv3n, tailn] if t]

        narrowed = []
        for r in rows:
            try:
                txt = N(r.text).replace("\n", " ")
                if all(tok in txt for tok in must_tokens):
                    narrowed.append(r)
            except Exception:
                pass

        return narrowed

    @pyqtSlot(object)
    def process(self, targets):
        if not self.is_ready():
            self.processed.emit([{'row': t.get('row'), 'ok': False, 'msg': '세션 준비 안됨'} for t in (targets or [])])
            return
        if self._busy:
            self.processed.emit([{'row': t.get('row'), 'ok': False, 'msg': '다른 작업 실행중'} for t in (targets or [])])
            return

        self._busy = True
        self.busyChanged.emit(True)
        self._cancel = False

        self._check_cancel("process")

        results = []
        try:
            d = self.driver

            def count_user_rows(user_text: str) -> int:
                self._go_iframe()
                user_eq = f".//td[normalize-space()='{user_text}']"
                code_pred = f".//td[contains(normalize-space(), '{code5}')]"
                if code4:
                    code_pred = f"({code_pred} or .//td[contains(normalize-space(), '{code4}')])"
                xpath = f"//tbody/tr[{user_eq} and {code_pred}]"
                rows = d.find_elements(By.XPATH, xpath)
                rows = self._narrow_by_path_strict(rows, lv2, lv3, path_hint)
                return len(rows)

            self._go_iframe()

            def _row_signature(tr):
                try:
                    tds = tr.find_elements(By.TAG_NAME, "td")
                    txts = []
                    for td in tds:
                        txt = (td.text or "").replace("\u00a0"," ").strip()
                        txts.append(re.sub(r"\s+", " ", txt))
                    return " | ".join(txts)
                except Exception:
                    return ""

            def _signature_exists(sig: str) -> bool:
                found = False
                def _find_on_page():
                    nonlocal found
                    self._go_iframe()
                    rows = d.find_elements(By.CSS_SELECTOR, "tbody tr")
                    for rnode in rows:
                        if _row_signature(rnode) == sig:
                            found = True
                            return True
                    return False
                self._for_each_page(_find_on_page)
                return found

            for t in (targets or []):
                self._check_cancel("process")
                row_idx = t.get('row')
                user = (t.get('user') or '').strip()
                lv2 = (t.get('lv2') or '').strip()
                lv3 = (t.get('lv3') or '').strip()
                path_hint = (t.get('path') or '').strip()
                kind = (t.get('kind') or '진행').strip()
                req = (t.get('req') or REQ_GRANT).strip()
                proj = (t.get('proj') or '').strip()
                code5, code4 = self._bus_code_variants(proj)

                ok = False
                msg = ""

                try:
                    if kind == "종료":
                        self._set_end_filter(END_FLAG_RELEASE if req == REQ_RELEASE else END_FLAG_GRANT)
                    else:
                        self._set_request_filter(req)

                    self._click_search()
                    time.sleep(0.6)
                    self._goto_first_page()

                    self._stabilize_grid()

                    before_total = 0
                    def sum_on_page():
                        nonlocal before_total
                        before_total += count_user_rows(user)
                        return False
                    self._for_each_page(sum_on_page)
                    target_sig_holder = {"sig": ""}

                    def work_on_page():
                        self._check_cancel("process")
                        self._go_iframe()

                        user_eq = f".//td[normalize-space()='{user}']"
                        code_pred = f".//td[contains(normalize-space(), '{code5}')]"
                        if code4:
                            code_pred = f"({code_pred} or .//td[contains(normalize-space(), '{code4}')])"

                        xpath = f"//tbody/tr[{user_eq} and {code_pred}]"
                        cand_rows = d.find_elements(By.XPATH, xpath)
                        
                        if (t.get('kind') or '').strip() != '종료':
                            cand_rows = self._narrow_by_path_strict(cand_rows, lv2, lv3, path_hint)
                       
                        if not cand_rows:
                            self._debug_dump("no_rows_after_filter")
                            return False

                        target_row = cand_rows[0]

                        req_is_release = (t.get('req') or REQ_GRANT) == REQ_RELEASE
                        if req_is_release:
                            btn_xpath = (
                                ".//button[normalize-space()='해제처리' "
                                "or contains(@class,'btn') and (contains(normalize-space(),'해제') or contains(@onclick,'Release'))]"
                            )
                        else:
                            btn_xpath = (
                                ".//button[normalize-space()='완료처리' "
                                "or normalize-space()='처리' "
                                "or contains(@class,'btn-outline-aurora')]"
                            )

                        target_sig_holder["sig"] = _row_signature(target_row)

                        self._check_cancel("process")

                        try:
                            d.execute_script("arguments[0].scrollIntoView({block:'center'});", target_row)
                        except Exception:
                            pass

                        def _btn_ready(_):
                            try:
                                el = target_row.find_element(By.XPATH, btn_xpath)
                                return el if el.is_displayed() and el.is_enabled() else False
                            except StaleElementReferenceException:
                                return False

                        try:
                            btn = WebDriverWait(target_row, 10).until(_btn_ready)
                        except TimeoutException:
                            self._debug_dump("no_action_button_wait_timeout")
                            return False

                        try:
                            d.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                        except Exception:
                            pass

                        try:
                            btn.click()
                        except ElementClickInterceptedException:
                            d.execute_script("arguments[0].click();", btn)
                        except Exception:
                            pass

                        if self._cancel:
                            try:
                                self._wait_overlay_gone(0.2)
                                try:
                                    cancel_btn = WebDriverWait(d, 1).until(
                                        EC.element_to_be_clickable((By.CSS_SELECTOR, "button.swal2-cancel"))
                                    )
                                    try:
                                        cancel_btn.click()
                                    except ElementClickInterceptedException:
                                        d.execute_script("arguments[0].click();", cancel_btn)
                                except TimeoutException:
                                    d.execute_script("if (window.Swal) { try{Swal.close();}catch(e){} }")
                            except Exception:
                                pass
                            self._check_cancel("process")
                            return False

                        try:
                            WebDriverWait(d, 3).until(
                                EC.visibility_of_element_located((By.CSS_SELECTOR, ".swal2-container"))
                            )
                            self._check_cancel("process")
                            if self._cancel:
                                try:
                                    cancel_btn = WebDriverWait(d, 1).until(
                                        EC.element_to_be_clickable((By.CSS_SELECTOR, "button.swal2-cancel"))
                                    )
                                    try:
                                        cancel_btn.click()
                                    except ElementClickInterceptedException:
                                        d.execute_script("arguments[0].click();", cancel_btn)
                                except TimeoutException:
                                    d.execute_script("if (window.Swal) { try{Swal.close();}catch(e){} }")
                                self._check_cancel("process")
                                return False

                            okbtn = WebDriverWait(d, 5).until(
                                EC.element_to_be_clickable((By.CSS_SELECTOR, "button.swal2-confirm"))
                            )
                            try:
                                okbtn.click()
                            except ElementClickInterceptedException:
                                d.execute_script("arguments[0].click();", okbtn)
                        except TimeoutException:
                            pass
                        
                        self._wait_overlay_gone(10)
                        self._stabilize_grid()                        
                        return True

                    self._goto_first_page()
                    did_click = self._for_each_page(work_on_page)
                    if not did_click:
                        ok, msg = False, "사번/프로젝트코드 불일치"
                    else:
                        timeout = time.time() + 12
                        disappeared = False
                        while time.time() < timeout:
                            self._check_cancel("process")
                            self._click_search()
                            time.sleep(0.8)
                            if target_sig_holder["sig"] and not _signature_exists(target_sig_holder["sig"]):
                                disappeared = True
                                break

                        if disappeared:
                            ok, msg = True, "처리 완료"
                        else:
                            after_total = 0
                            def sum_after_page():
                                nonlocal after_total
                                after_total += count_user_rows(user)
                                return False
                            self._for_each_page(sum_after_page)

                            ok = after_total < before_total
                            msg = f"처리 {'완료' if ok else '미확인'}: 전={before_total}, 후={after_total}"

                except SystemExit:
                    return
                except Exception as e:
                    ok, msg = False, f"오류: {e}"

                if not ok and self.debug_enabled:
                    try:
                        ts = int(time.time())
                        tag_user = re.sub(r'[^0-9A-Za-z_-]+', '_', user)[:30]
                        tag_lv3  = re.sub(r'[^0-9A-Za-z_-]+', '_', lv3)[:20]
                        base = f"fail_{ts}_{tag_user}_{tag_lv3}"
                        with open(os.path.join(self.debug_dir, f"{base}.html"), "w", encoding="utf-8") as f:
                            f.write(d.page_source)
                        try:
                            d.save_screenshot(os.path.join(self.debug_dir, f"{base}.png"))
                        except Exception:
                            pass
                    except Exception:
                        pass

                results.append({'row': row_idx, 'ok': ok, 'msg': msg})

            self._busy = False
            self.busyChanged.emit(False)
            self.processed.emit(results)

        except SystemExit:
            self._busy = False
            self.busyChanged.emit(False)
            if results:
                self.processed.emit(results)
            return
        except Exception as e:
            self._busy = False
            self.busyChanged.emit(False)
            self.processed.emit([{'row': t.get('row'), 'ok': False, 'msg': f'오류: {e}'} for t in (targets or [])])

class BusWatcher(QObject):
    readyChanged = pyqtSignal(bool, str)
    countsReady = pyqtSignal(dict)

    def __init__(self, dl_dir):
        super().__init__()
        self._user = ""
        self._pw = ""
        self._last_counts = {"신규미완료":0, "진행-부여":0, "진행-제거":0, "종료-부여":0, "종료-제거":0}
        self._mgr = BusSessionManager(dl_dir)

    def set_creds(self, user, pw):
        self._user, self._pw = (user or "").strip(), (pw or "").strip()

    def set_debug(self, on, debug_dir):
        try:
            self._mgr.set_debug(bool(on), debug_dir)
        except Exception:
            pass

    def is_ready(self):
        try:
            return self._mgr.is_ready()
        except Exception:
            return False

    @pyqtSlot()
    def start(self):
        if not (self._user and self._pw):
            self.readyChanged.emit(False, "워처 자격증명 없음")
            return
        try:
            self._mgr.set_creds(self._user, self._pw)
            self._mgr.start()
            self.readyChanged.emit(True, "워처 준비됨")
        except Exception:
            self.readyChanged.emit(False, "워처 시작 실패")

    @pyqtSlot()
    def stop(self):
        try:
            self._mgr.stop()
        except Exception:
            pass
        self.readyChanged.emit(False, "워처 정지됨")

    @pyqtSlot()
    def collect_counts(self):
        try:
            self._mgr.collect_request_counts()
            self._last_counts = getattr(self._mgr, "_last_counts", self._last_counts)
            self.countsReady.emit(self._last_counts)
        except Exception:
            self.countsReady.emit(self._last_counts)

class CheckBoxHeader(QHeaderView):
    stateChanged = pyqtSignal(bool)

    def __init__(self, orientation, parent=None):
        super().__init__(orientation, parent)
        self.setSectionsClickable(True)
        self.setHighlightSections(False)

        self._cb = QCheckBox(self)
        self._cb.setTristate(False)
        self._cb.setChecked(True)
        self._cb.setStyleSheet("QCheckBox{background:transparent; padding:0; margin:0;}")

        self._cb.stateChanged.connect(lambda st: self.stateChanged.emit(st == Qt.Checked))

        self.sectionResized.connect(self._reposition)
        self.sectionMoved.connect(self._reposition)
        self.geometriesChanged.connect(self._reposition)
        self.sectionCountChanged.connect(lambda *_: self._reposition())

        self._reposition()

    def setChecked(self, on: bool):
        self._cb.blockSignals(True)
        self._cb.setChecked(bool(on))
        self._cb.blockSignals(False)
        self.updateSection(0)

    def _reposition(self, *args):
        if self.count() == 0:
            return

        x0 = self.sectionViewportPosition(0)
        w0 = self.sectionSize(0)
        h  = self.height()

        sz = self._cb.sizeHint()
        x  = x0 + max(0, (w0 - sz.width()) // 2)
        y  = max(0, (h  - sz.height()) // 2)

        if w0 < sz.width():
            x = x0 + 2

        self._cb.setGeometry(x, y, sz.width(), sz.height())
        self._cb.raise_()
        self._cb.show()

    def mousePressEvent(self, e):
        if self.logicalIndexAt(e.pos()) == 0:
            self._cb.toggle()
            e.accept()
            return
        super().mousePressEvent(e)


class CopyTable(QTableWidget):
    def keyPressEvent(self, event):
        if event.matches(QKeySequence.Copy):
            ranges = self.selectedRanges()
            if not ranges:
                super().keyPressEvent(event); return
            blocks = []
            for rng in ranges:
                rows = []
                for r in range(rng.topRow(), rng.bottomRow()+1):
                    cols = []
                    for c in range(rng.leftColumn(), rng.rightColumn()+1):
                        it = self.item(r, c)
                        cols.append("" if it is None else it.text())
                    rows.append("\t".join(cols))
                blocks.append("\n".join(rows))
            QGuiApplication.clipboard().setText("\n\n".join(blocks))
        else:
            super().keyPressEvent(event)

class ManualEntryDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("수동 입력")
        self.setModal(True)

        self.cb_kind = QComboBox()
        self.cb_kind.addItems(["진행", "종료"])
        self.cb_kind.setCurrentText("진행")
        self.cb_kind.currentTextChanged.connect(self._on_kind_changed)

        self.cb_reqtype = QComboBox()
        self.le_user = QLineEdit()
        self.le_proj = QLineEdit()
        self.cb_lv2  = QComboBox()
        self.cb_lv3  = QComboBox()
        self.cb_role = QComboBox()
        
        self.cb_reqtype.addItems([REQ_GRANT, REQ_RELEASE])
        self.cb_reqtype.setCurrentText(REQ_GRANT)

        self.cb_lv2.addItems(["Study", "Isolated"])
        self.cb_lv2.setCurrentIndex(-1)

        self.cb_lv3.addItems(LEVEL3_CHOICES)
        self.cb_lv3.setCurrentIndex(-1)

        self.cb_role.addItem("")
        self.cb_role.setEnabled(False)

        self.cb_lv2.currentTextChanged.connect(self._refresh_role_candidates)
        self.cb_lv3.currentTextChanged.connect(self._refresh_role_candidates)
        self.le_proj.textChanged.connect(self._refresh_role_candidates)

        form = QFormLayout()
        form.addRow("구분*", self.cb_kind) 
        form.addRow("요청사항*", self.cb_reqtype) 
        form.addRow("사번*", self.le_user)
        form.addRow("프로젝트코드*", self.le_proj)
        form.addRow("Level2*", self.cb_lv2)
        form.addRow("Level3*", self.cb_lv3)
        form.addRow("STATROLE", self.cb_role)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, parent=self)
        btns.accepted.connect(self._on_ok)
        btns.rejected.connect(self.reject)

        lay = QVBoxLayout(self)
        lay.addLayout(form)
        lay.addWidget(btns)

        self.result_row = None

    def _is_new_by_proj_text(self) -> bool:
        txt = self.le_proj.text().strip()
        if not txt:
            return True
        try:
            return is_new_template(txt)
        except Exception:
            return True

    def _refresh_role_candidates(self):
        try:
            lv2 = self.cb_lv2.currentText().strip() if self.cb_lv2.currentIndex() != -1 else ""
            lv3 = self.cb_lv3.currentText().strip() if self.cb_lv3.currentIndex() != -1 else ""
            is_new = self._is_new_by_proj_text()

            self.cb_role.blockSignals(True)
            self.cb_role.clear()
            self.cb_role.addItem("")
            self.cb_role.blockSignals(False)
            self.cb_role.setCurrentIndex(0)
            self.cb_role.setEnabled(False)

            if not is_stat_lv3(lv3):
                return

            candidates = []
            if lv2 == "Study":
                candidates = sorted(STUDY_ROLES)
            elif lv2 == "Isolated":
                if is_stat_idmc_lv3(lv3):
                    if is_stat_idmc_new_policy(self.le_proj.text().strip()):
                        candidates = sorted(STUDY_ROLES)
                else:
                    candidates = sorted(ISOLATED_ROLES) if is_new else ["Randomization Statistician"]

            if candidates:
                self.cb_role.blockSignals(True)
                for r in candidates:
                    self.cb_role.addItem(r)
                self.cb_role.blockSignals(False)
                self.cb_role.setEnabled(True)
        except Exception as e:
            parent = self.parent()
            if parent and hasattr(parent, "_log"):
                parent._log(f"[역할후보 갱신 예외] {e}")

    def _on_kind_changed(self, text: str):
        """구분이 '종료'면 Level2/Level3 입력을 비활성화하고 값을 비움."""
        is_end = (text or "").strip() == "종료"

        for cb in (self.cb_lv2, self.cb_lv3):
            try:
                cb.blockSignals(True)
                if is_end:
                    cb.setCurrentIndex(-1)
                cb.setEnabled(not is_end)
            finally:
                cb.blockSignals(False)

        try:
            self._refresh_role_candidates()
        except Exception:
            pass

    def _on_ok(self):
        try:
            kind   = self.cb_kind.currentText().strip()
            reqtype = self.cb_reqtype.currentText().strip()
            user = self.le_user.text().strip()
            proj = self.le_proj.text().strip()
            lv2  = self.cb_lv2.currentText().strip() if self.cb_lv2.currentIndex() != -1 else ""
            lv3  = self.cb_lv3.currentText().strip() if self.cb_lv3.currentIndex() != -1 else ""
            role = self.cb_role.currentText().strip() if self.cb_role.isEnabled() else ""

            missing = []
            if not user: missing.append("사번")
            if not proj: missing.append("프로젝트코드")
            if kind != "종료":
                if not lv2:  missing.append("Level2")
                if not lv3:  missing.append("Level3")
            if missing:
                QMessageBox.warning(self, "입력 누락", f"다음 항목을 입력해 주세요: {', '.join(missing)}")
                return

            if not is_stat_lv3(lv3):
                role = ""

            self.result_row = (reqtype, user, proj, lv2, lv3, "", role, kind)
            self.accept()
        except Exception as e:
            try:
                parent = self.parent()
                if parent and hasattr(parent, "_log"):
                    parent._log(f"[수동입력 예외] {e}")
            except:
                pass
            QMessageBox.critical(self, "오류", f"수동 입력 처리 중 오류:\n{e}")

    def _set_roles_for_lv2(self, lv2_text):
        self.cb_role.clear()
        self.cb_role.addItem("")
        if lv2_text == "Study":
            for r in sorted(STUDY_ROLES):
                self.cb_role.addItem(r)
        elif lv2_text == "Isolated":
            for r in sorted(ISOLATED_ROLES):
                self.cb_role.addItem(r)
        self.cb_role.setCurrentIndex(0)

class CheckBoxHeaderAt(QHeaderView):
    stateChanged = pyqtSignal(bool)

    def __init__(self, orientation, parent=None, target_index: int = 0, initially_checked: bool = True):
        super().__init__(orientation, parent)
        self._target = int(target_index)

        self._cb = QCheckBox(self)
        self._cb.setTristate(False)
        self._cb.setChecked(bool(initially_checked))
        self._cb.setStyleSheet("QCheckBox{background:transparent; padding:0; margin:0;}")
        self._cb.stateChanged.connect(lambda st: self.stateChanged.emit(st == Qt.Checked))

        self.setSectionsClickable(True)
        self.setHighlightSections(False)

        self.sectionResized.connect(self._reposition)
        self.sectionMoved.connect(self._reposition)
        self.sectionCountChanged.connect(lambda *_: self._reposition())

    def showEvent(self, e):
        super().showEvent(e)
        self._reposition()

    def setTargetSection(self, idx: int):
        self._target = int(idx)
        self._reposition()

    def setChecked(self, on: bool):
        self._cb.blockSignals(True)
        self._cb.setChecked(bool(on))
        self._cb.blockSignals(False)
        self.updateSection(self._target)

    def _reposition(self, *args):
        if self.count() == 0 or self._target < 0 or self._target >= self.count():
            self._cb.hide(); return
        x0 = self.sectionViewportPosition(self._target)
        w0 = self.sectionSize(self._target)
        h  = self.height()
        sz = self._cb.sizeHint()
        x  = x0 + max(0, (w0 - sz.width()) // 2)
        y  = max(0, (h  - sz.height()) // 2)
        self._cb.setGeometry(x, y, sz.width(), sz.height())
        self._cb.raise_()
        self._cb.show()

    def mousePressEvent(self, e):
        if self.logicalIndexAt(e.pos()) == self._target:
            self._cb.toggle()
            e.accept()
            return
        super().mousePressEvent(e)

class ManualNewRequestDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("수동 추가")
        lay = QVBoxLayout(self)

        form = QFormLayout()
        self.edt_proj = QLineEdit(self)
        self.edt_proj.setPlaceholderText("예: 25001")
        self.edt_proj.setMaxLength(5)
        self.edt_name = QLineEdit(self)
        self.edt_name.setPlaceholderText("프로젝트명")
        form.addRow("프로젝트 코드", self.edt_proj)
        form.addRow("프로젝트명", self.edt_name)
        lay.addLayout(form)

        btn_box = QDialogButtonBox(Qt.Horizontal, self)
        self.btn_ok = btn_box.addButton("생성", QDialogButtonBox.AcceptRole)
        self.btn_cancel = btn_box.addButton("취소", QDialogButtonBox.RejectRole)
        self.btn_ok.clicked.connect(self.accept)
        self.btn_cancel.clicked.connect(self.reject)
        lay.addWidget(btn_box)

    def values(self) -> tuple[str, str]:
        return self.edt_proj.text().strip(), self.edt_name.text().strip()


class NewItemsViewer(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("신규 폴더 생성요청 목록")
        self.resize(800, 520)

        lay = QVBoxLayout(self)
        self.tbl = QTableWidget(0, 0, self)

        hdr = CheckBoxHeader(Qt.Horizontal, self.tbl)
        hdr.stateChanged.connect(self._toggle_all_rows)
        self.tbl.setHorizontalHeader(hdr)

        self.tbl.setAlternatingRowColors(True)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl.horizontalHeader().setStretchLastSection(True)

        btns = QHBoxLayout()
        self.btn_manual = QPushButton("수동 추가")
        self.btn_create = QPushButton("자동 생성")
        self.btn_refresh = QPushButton("↻ 새로고침")
        self.btn_close = QPushButton("닫기")
        self.btn_close.clicked.connect(self.accept)
        self.btn_create.clicked.connect(self._on_create_clicked)
        self.btn_manual.clicked.connect(self._on_manual_clicked)
        btns.addStretch()
        btns.addWidget(self.btn_manual)
        btns.addWidget(self.btn_create)
        btns.addWidget(self.btn_refresh)
        btns.addWidget(self.btn_close)
        lay.addWidget(self.tbl); lay.addLayout(btns)

        self.status_lbl = QLabel("")
        self.prg = QProgressBar(self)
        self.prg.setRange(0, 0)
        self.prg.hide()

        bar = QHBoxLayout()
        bar.addWidget(self.status_lbl)
        bar.addStretch()
        bar.addWidget(self.prg)
        lay.addLayout(bar)

        self._last_worker_msg = ""

    def _set_busy(self, on: bool, msg: str = ""):
        self.status_lbl.setText(msg or "")
        self.prg.setVisible(on)
        self.btn_manual.setEnabled(not on)
        self.btn_create.setEnabled(not on)
        self.btn_refresh.setEnabled(not on)
        self.btn_close.setEnabled(not on)
        QApplication.processEvents()

    def _on_worker_progress(self, cur: int, total: int, message: str):
        self.status_lbl.setText(message)
        self._last_worker_msg = message
        msg = (message or "").strip()
        if msg.startswith("[REQUEST_READ]"):
            append_request_trace_log("request_read", msg)
        elif msg.startswith("[REQUEST_PERF]"):
            append_request_trace_log("request_perf", msg)
        QApplication.processEvents()

    def _on_worker_finished(self, ok_cnt: int, fail_cnt: int):
        base = f"완료: 성공 {ok_cnt} / 실패 {fail_cnt}"
        if fail_cnt > 0 and self._last_worker_msg:
            base += f"\n마지막 메시지: {self._last_worker_msg}"

        self._set_busy(False, base)
        try:
            self.worker_thread.quit()
            self.worker_thread.wait(1500)
        except Exception:
            pass

        self.worker.deleteLater()
        self.worker_thread.deleteLater()
        self.worker = None
        self.worker_thread = None

    def _manual_add(self, proj: str, name: str):
        parent = self.parent()
        ps_path = getattr(parent, "ps_path", shutil.which("pwsh") or shutil.which("powershell") or "powershell")
        ps_kind = getattr(parent, "ps_kind", "pwsh" if "pwsh" in os.path.basename(ps_path).lower() else "powershell")

        worker = CreateWorker([], {}, ps_path, ps_kind)
        self._set_busy(True, f"{proj} 폴더/권한 생성 중…")
        ok, msg = worker._create_group_and_folder(proj, name)
        self._set_busy(False, "")

        if ok:
            QMessageBox.information(self, "완료", f"{proj} 생성 완료")
        else:
            QMessageBox.critical(self, "오류", f"{proj} 생성 실패: {msg}")

    def _on_manual_clicked(self):
        dlg = ManualNewRequestDialog(self)
        if dlg.exec_() != QDialog.Accepted:
            return

        proj, name = dlg.values()
        if not proj:
            QMessageBox.warning(self, "알림", "프로젝트 코드를 입력하세요.")
            return

        self._manual_add(proj, name)

    def _hide_empty_columns(self, always_show: set[str] | None = None, skip_col_idx: int | None = None):
        t = self.tbl
        if t is None:
            return
        rows, cols = t.rowCount(), t.columnCount()
        always_show = always_show or set()

        for c in range(cols):
            t.setColumnHidden(c, False)

        for c in range(cols):
            if skip_col_idx is not None and c == skip_col_idx:
                continue
            header = t.horizontalHeaderItem(c).text() if t.horizontalHeaderItem(c) else ""
            if header in always_show:
                continue
            has_data = False
            for r in range(rows):
                it = t.item(r, c)
                if it and it.text().strip():
                    has_data = True
                    break
            t.setColumnHidden(c, not has_data)

    def set_data(self, header, rows):
        try:
            self.tbl.blockSignals(True)

            col_select = 0
            col_count = 1 + len(header)

            self.tbl.clear()
            self.tbl.setColumnCount(col_count)
            self.tbl.setRowCount(len(rows))

            it_sel = QTableWidgetItem("")
            it_sel.setToolTip("선택")
            self.tbl.setHorizontalHeaderItem(col_select, it_sel)
            self.tbl.horizontalHeader().setSectionResizeMode(col_select, QHeaderView.Fixed)
            self.tbl.setColumnWidth(col_select, 28)

            for i, h in enumerate(header, start=1):
                self.tbl.setHorizontalHeaderItem(i, QTableWidgetItem(h or ""))

            for r, row in enumerate(rows):
                cb = QCheckBox(self.tbl)
                cb.setChecked(True)
                wrap = QWidget(self.tbl)
                box = QHBoxLayout(wrap); box.setContentsMargins(0,0,0,0)
                box.addWidget(cb, 0, Qt.AlignCenter)
                self.tbl.setCellWidget(r, col_select, wrap)

                for c, v in enumerate(row, start=1):
                    it = QTableWidgetItem(str(v) if v is not None else "")
                    it.setTextAlignment(Qt.AlignCenter)
                    self.tbl.setItem(r, c, it)

            for c in range(1, col_count):
                self.tbl.horizontalHeader().setSectionResizeMode(c, QHeaderView.ResizeToContents)
            self.tbl.horizontalHeader().setStretchLastSection(True)

            self._hide_empty_columns(always_show={"프로젝트코드","프로젝트명"}, skip_col_idx=0)
        finally:
            self.tbl.blockSignals(False)

        self._hidx = { (self.tbl.horizontalHeaderItem(i).text() or ""): i
                       for i in range(1, self.tbl.columnCount()) }

    def _toggle_all_rows(self, checked: bool):
        t = self.tbl
        for r in range(t.rowCount()):
            w = t.cellWidget(r, 0)
            if not w:
                continue
            cb = w.findChild(QCheckBox)
            if cb:
                cb.setChecked(checked)

    def _selected_targets(self):
        t = self.tbl
        items = []
        code_col = self._hidx.get("프로젝트코드", -1)
        name_col = self._hidx.get("프로젝트명", -1)

        if code_col < 0:
            QMessageBox.warning(self, "알림", "헤더 '프로젝트코드'를 찾을 수 없습니다.")
            return items

        for r in range(t.rowCount()):
            w = t.cellWidget(r, 0)
            cb = w.findChild(QCheckBox) if w else None
            if not (cb and cb.isChecked()):
                continue
            proj = (t.item(r, code_col).text().strip() if t.item(r, code_col) else "")
            pname = (t.item(r, name_col).text().strip() if (name_col >= 0 and t.item(r, name_col)) else "")
            if proj:
                items.append((proj, pname))
        return items

    def _run_pwsh(self, cmd: str) -> tuple[bool, str]:
        try:
            parent = self.parent()
            if parent and hasattr(parent, "_wrap_cmd_utf8") and hasattr(parent, "_pick_powershell"):
                wrapped = parent._wrap_cmd_utf8(cmd)
                ps_path, _ = parent._pick_powershell()
            else:
                wrapped = cmd
                ps_path = shutil.which("pwsh") or shutil.which("powershell") or "powershell"

            p = QProcess(self)
            args = ["-NoLogo","-NoProfile","-ExecutionPolicy","Bypass","-Command", wrapped]
            p.start(ps_path, args)
            if not p.waitForStarted(30000):
                return False, "PowerShell 시작 실패"

            if not p.waitForFinished(300000):
                try:
                    p.kill(); p.waitForFinished(3000)
                except Exception:
                    pass
                return False, "PowerShell 실행 시간 초과"

            rc  = p.exitCode()
            out = bytes(p.readAllStandardOutput()).decode("utf-8","ignore").strip()
            err = bytes(p.readAllStandardError()).decode("utf-8","ignore").strip()
            return (rc == 0), (err or out)
        except Exception as e:
            return False, f"예외: {e}"

    def _on_create_clicked(self):
        items = []
        t = self.tbl
        proj_col = None; name_col = None
        for c in range(1, t.columnCount()):
            h = t.horizontalHeaderItem(c)
            ht = (h.text() if h else "").strip()
            if ht == "프로젝트코드": proj_col = c
            if ht in ("프로젝트명", "과제명", "제목"): name_col = c
        if proj_col is None:
            QMessageBox.warning(self, "알림", "프로젝트코드 컬럼을 찾을 수 없습니다."); return

        for r in range(t.rowCount()):
            w = t.cellWidget(r, 0)
            if not w: continue
            cb = w.findChild(QCheckBox)
            if not (cb and cb.isChecked()): 
                continue
            proj = (t.item(r, proj_col).text().strip() if t.item(r, proj_col) else "")
            name = (t.item(r, name_col).text().strip() if (name_col is not None and t.item(r, name_col)) else "")
            if proj:
                items.append({"proj": proj, "name": name})

        if not items:
            QMessageBox.information(self, "알림", "선택된 항목이 없습니다."); return

        parent = self.parent()
        creds = getattr(parent, "creds", {}) if parent else {}
        if not creds.get("id") or not creds.get("pw"):
            QMessageBox.warning(self, "알림", "BUS 계정이 필요합니다. 설정에서 저장 후 다시 시도하세요.")
            return

        ps_path = getattr(parent, "ps_path", shutil.which("pwsh") or shutil.which("powershell") or "powershell")
        ps_kind = getattr(parent, "ps_kind", "pwsh" if "pwsh" in os.path.basename(ps_path).lower() else "powershell")

        self.worker_thread = QThread(self)
        self.worker = CreateWorker(items, creds, ps_path, ps_kind)
        self.worker.moveToThread(self.worker_thread)

        self.worker.progress.connect(self._on_worker_progress)
        self.worker.finished.connect(self._on_worker_finished)
        self.worker.error.connect(lambda msg: QMessageBox.critical(self, "오류", msg))

        self.worker_thread.started.connect(self.worker.run)

        self._set_busy(True, f"처리 시작… (대상 {len(items)}건)")
        self.prg.setRange(0, 0)
        self.worker_thread.start()

class CreateWorker(QObject):
    progress = pyqtSignal(int, int, str)
    finished = pyqtSignal(int, int)
    error = pyqtSignal(str)

    def __init__(self, items, creds, ps_path, ps_kind, parent=None):
        super().__init__(parent)
        self.items = items
        self.creds = creds
        self.ps_path = ps_path
        self.ps_kind = ps_kind
        self._stop = False

    def stop(self):
        self._stop = True

    def _wrap_cmd_utf8(self, cmd: str) -> str:
        pre = "$OutputEncoding=[System.Text.Encoding]::UTF8; [Console]::OutputEncoding=[System.Text.Encoding]::UTF8;"
        if self.ps_kind == "powershell":
            pre += " chcp 65001 > $null;"
        return pre + " " + cmd

    def _run_pwsh(self, cmd: str, timeout_ms=180000) -> (bool, str):
        from PyQt5.QtCore import QProcess
        p = QProcess()
        p.start(self.ps_path, ["-NoLogo","-NoProfile","-ExecutionPolicy","Bypass","-Command", self._wrap_cmd_utf8(cmd)])
        if not p.waitForStarted(15000):
            return False, "PowerShell 시작 실패"

        deadline = time.time() + (timeout_ms/1000.0)
        ok = False
        while time.time() < deadline:
            if self._stop:
                try:
                    p.kill()
                    p.waitForFinished(3000)
                except Exception:
                    pass
                return False, "사용자 중지"
            QApplication.processEvents()
            if p.waitForFinished(50):
                ok = (p.exitCode() == 0)
                break
        if not ok and p.state() == QProcess.Running:
            try:
                p.kill()
                p.waitForFinished(3000)
            except Exception:
                pass
            return False, "PowerShell 실행 시간 초과"

        out = bytes(p.readAllStandardOutput()).decode("utf-8","ignore")
        err = bytes(p.readAllStandardError()).decode("utf-8","ignore")
        return ok, (err.strip() or out.strip())

    def _bus_click_process(self, proj_code: str) -> (bool, str):
        d = None
        stage = "INIT"
        bus_msg_title = ""
        bus_msg_body = ""
        read_mode = (self.creds.get("request_read_mode") or "observe_only").strip().lower()
        if read_mode not in ("legacy_only", "observe_only", "state_driven"):
            read_mode = "observe_only"
        try:
            options = webdriver.ChromeOptions()
            options.add_argument("--headless=new")
            options.add_argument("--disable-gpu")
            options.add_argument("--no-sandbox")
            options.add_argument("--window-size=1280,900")
            d = webdriver.Chrome(options=options)

            def normalize_text(s: str) -> str:
                s = (s or "").replace("\u00a0", " ").strip()
                return re.sub(r"\s+", " ", s)

            def set_select_value_and_fire(driver, select_id, value):
                driver.execute_script("""
                    var s = document.getElementById(arguments[0]);
                    if (s) {
                        s.value = arguments[1];
                        try { s.dispatchEvent(new Event('change', {bubbles:true})); } catch(e) {}
                        if (typeof AllBtnYn === 'function') { try { AllBtnYn(); } catch(e) {} }
                    }
                """, select_id, value)

            def click_search_manual(driver):
                try:
                    WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.ID, "btnSearch"))
                    ).click()
                except Exception:
                    driver.switch_to.default_content()
                    WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.ID, "btnSearch"))
                    ).click()
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_all_elements_located((By.TAG_NAME, "iframe"))
                    )
                    driver.switch_to.frame(driver.find_elements(By.TAG_NAME, "iframe")[0])

            def find_swal_popup(driver):
                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass

                popups = driver.find_elements(By.CSS_SELECTOR, "div.swal2-container.swal2-shown")
                if popups:
                    return popups[0]

                try:
                    frames = driver.find_elements(By.TAG_NAME, "iframe")
                except Exception:
                    frames = []

                for fr in frames:
                    try:
                        driver.switch_to.frame(fr)
                        popups = driver.find_elements(By.CSS_SELECTOR, "div.swal2-container.swal2-shown")
                        if popups:
                            return popups[0]
                    except Exception:
                        pass
                    try:
                        driver.switch_to.default_content()
                    except Exception:
                        pass

                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass
                return None

            def enter_first_iframe(driver, timeout=15, reload_url=None):
                for attempt in range(2):
                    try:
                        driver.switch_to.default_content()
                    except Exception:
                        pass

                    try:
                        WebDriverWait(driver, timeout).until(
                            lambda drv: len(drv.find_elements(By.TAG_NAME, "iframe")) > 0
                        )
                        frames = driver.find_elements(By.TAG_NAME, "iframe")
                        if frames:
                            driver.switch_to.frame(frames[0])
                            return True
                    except Exception:
                        pass

                    if reload_url and attempt == 0:
                        try:
                            driver.get(reload_url)
                        except Exception:
                            pass
                        time.sleep(0.5)

                return False

            def refocus_first_iframe(driver, reload_url=None):
                return enter_first_iframe(driver, timeout=5, reload_url=reload_url)

            def _fmt_ts():
                return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            def is_loading_state(driver):
                out = {"loading": False, "reason": "no loading marker", "exception": None}
                try:
                    markers = [
                        (By.CSS_SELECTOR, ".loading"),
                        (By.CSS_SELECTOR, ".blockUI"),
                        (By.CSS_SELECTOR, ".blockOverlay"),
                        (By.CSS_SELECTOR, ".swal2-container.swal2-shown"),
                    ]
                    hits = []
                    for by, sel in markers:
                        nodes = driver.find_elements(by, sel)
                        vis = [n for n in nodes if n.is_displayed()]
                        if vis:
                            hits.append(sel)
                    if hits:
                        out["loading"] = True
                        out["reason"] = "visible: " + ",".join(hits)
                except Exception as e:
                    out["exception"] = str(e)
                    out["reason"] = "loading check failed"
                return out

            def get_request_row_count(driver):
                out = {"row_count": 0, "table_found": False, "exception": None}
                try:
                    rows = driver.find_elements(By.CSS_SELECTOR, "tbody tr")
                    out["table_found"] = True
                    count = 0
                    for tr in rows:
                        try:
                            if "dataTables_empty" in (tr.get_attribute("class") or ""):
                                continue
                        except Exception:
                            pass
                        count += 1
                    out["row_count"] = count
                except Exception as e:
                    out["exception"] = str(e)
                return out

            def has_empty_list_message(driver):
                out = {
                    "empty_message": False,
                    "matched_selector": "",
                    "matched_text": "",
                    "exception": None,
                }
                try:
                    empty_cells = driver.find_elements(By.CSS_SELECTOR, "tbody td.dataTables_empty")
                    for el in empty_cells:
                        txt = normalize_text(el.text)
                        if txt:
                            out["empty_message"] = True
                            out["matched_selector"] = "tbody td.dataTables_empty"
                            out["matched_text"] = txt
                            return out

                    info_els = driver.find_elements(By.CSS_SELECTOR, ".dataTables_info")
                    if info_els:
                        txt = normalize_text(info_els[0].text)
                        if txt and (" 0 " in f" {txt} " or "0건" in txt or "없" in txt):
                            out["empty_message"] = True
                            out["matched_selector"] = ".dataTables_info"
                            out["matched_text"] = txt
                except Exception as e:
                    out["exception"] = str(e)
                return out

            def read_request_list_state(driver, page_name="new_request_list"):
                ts0 = time.perf_counter()
                result = {
                    "timestamp": _fmt_ts(),
                    "page_name": page_name,
                    "loading": False,
                    "row_count": 0,
                    "empty_message": False,
                    "state": "UNCERTAIN",
                    "reason": "not evaluated",
                    "exception": None,
                    "details": {},
                }
                perf = {
                    "cycle_id": int(time.time() * 1000),
                    "timestamp": result["timestamp"],
                    "refresh_ms": 0,
                    "loading_check_ms": 0,
                    "row_count_ms": 0,
                    "empty_msg_ms": 0,
                    "state_decision_ms": 0,
                    "total_ms": 0,
                    "final_state": "UNCERTAIN",
                }

                try:
                    t = time.perf_counter()
                    loading_res = is_loading_state(driver)
                    perf["loading_check_ms"] = int((time.perf_counter() - t) * 1000)

                    t = time.perf_counter()
                    row_res = get_request_row_count(driver)
                    perf["row_count_ms"] = int((time.perf_counter() - t) * 1000)

                    t = time.perf_counter()
                    empty_res = has_empty_list_message(driver)
                    perf["empty_msg_ms"] = int((time.perf_counter() - t) * 1000)

                    result["loading"] = bool(loading_res.get("loading"))
                    result["row_count"] = int(row_res.get("row_count") or 0)
                    result["empty_message"] = bool(empty_res.get("empty_message"))
                    result["details"] = {
                        "loading": loading_res,
                        "row_count": row_res,
                        "empty_message": empty_res,
                    }

                    t = time.perf_counter()
                    if loading_res.get("exception") or row_res.get("exception") or empty_res.get("exception"):
                        result["state"] = "ERROR"
                        result["reason"] = "reader exception in at least one stage"
                        result["exception"] = {
                            "loading": loading_res.get("exception"),
                            "row_count": row_res.get("exception"),
                            "empty_message": empty_res.get("exception"),
                        }
                    elif result["loading"]:
                        result["state"] = "LOADING"
                        result["reason"] = loading_res.get("reason") or "loading marker detected"
                    elif result["row_count"] >= 1:
                        result["state"] = "HAS_ITEMS"
                        result["reason"] = f"row_count={result['row_count']}"
                    elif result["row_count"] == 0 and result["empty_message"]:
                        result["state"] = "EMPTY_CONFIRMED"
                        result["reason"] = "loading finished and empty message detected"
                    else:
                        result["state"] = "UNCERTAIN"
                        result["reason"] = "loading finished but empty condition not clearly confirmed"
                    perf["state_decision_ms"] = int((time.perf_counter() - t) * 1000)
                    perf["final_state"] = result["state"]
                except Exception as e:
                    result["state"] = "ERROR"
                    result["reason"] = "unhandled read_request_list_state exception"
                    result["exception"] = str(e)
                    perf["final_state"] = "ERROR"
                finally:
                    perf["total_ms"] = int((time.perf_counter() - ts0) * 1000)
                return result, perf

            def emit_request_logs(read_result, perf_result, legacy_result):
                read_log = (
                    "[REQUEST_READ] "
                    f"time={read_result.get('timestamp')} "
                    f"page={read_result.get('page_name')} "
                    f"loading={read_result.get('loading')} "
                    f"row_count={read_result.get('row_count')} "
                    f"empty_message={read_result.get('empty_message')} "
                    f"state={read_result.get('state')} "
                    f"reason=\"{read_result.get('reason')}\" "
                    f"legacy_result={legacy_result} "
                    f"exception={read_result.get('exception')}"
                )
                self.progress.emit(0, 0, read_log)
                append_request_trace_log("request_read", read_log)

                perf_log = (
                    "[REQUEST_PERF] "
                    f"cycle_id={perf_result.get('cycle_id')} "
                    f"time={perf_result.get('timestamp')} "
                    f"refresh_ms={perf_result.get('refresh_ms')} "
                    f"loading_check_ms={perf_result.get('loading_check_ms')} "
                    f"row_count_ms={perf_result.get('row_count_ms')} "
                    f"empty_msg_ms={perf_result.get('empty_msg_ms')} "
                    f"state_decision_ms={perf_result.get('state_decision_ms')} "
                    f"total_ms={perf_result.get('total_ms')} "
                    f"final_state={perf_result.get('final_state')}"
                )
                self.progress.emit(0, 0, perf_log)
                append_request_trace_log("request_perf", perf_log)

            stage = "OPEN_LOGIN_PAGE"
            d.get(BUS_LOGIN_URL)
            WebDriverWait(d, 20).until(EC.presence_of_element_located((By.ID, "windowsaccount")))

            stage = "INPUT_IDPW"
            d.find_element(By.ID, "windowsaccount").clear()
            d.find_element(By.ID, "windowsaccount").send_keys(self.creds["id"])
            d.find_element(By.ID, "password").clear()
            d.find_element(By.ID, "password").send_keys(self.creds["pw"])

            stage = "CLICK_LOGIN"
            d.find_element(By.ID, "btnLogin").click()
            WebDriverWait(d, 20).until(EC.url_contains("Common"))

            stage = "OPEN_NEW_PAGE"
            refresh_t0 = time.perf_counter()
            d.get(BUS_NEW_URL)
            refresh_ms = int((time.perf_counter() - refresh_t0) * 1000)

            stage = "ENTER_IFRAME"
            if not enter_first_iframe(d, reload_url=BUS_NEW_URL):
                return False, f"IFRAME_ENTER_FAIL @ {stage}"

            try:
                pre_info = d.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip()
            except Exception:
                pre_info = ""

            stage = "SET_FILTER"
            try:
                set_select_value_and_fire(d, "processYn", "N")
            except Exception:
                pass

            stage = "CLICK_SEARCH"
            if not refocus_first_iframe(d, reload_url=BUS_NEW_URL):
                return False, f"IFRAME_LOST_BEFORE_SEARCH @ {stage}"
            click_search_manual(d)

            stage = "WAIT_FILTER_APPLY"
            try:
                WebDriverWait(d, 7).until(
                    lambda x: (
                        x.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip()
                        if x.find_elements(By.CSS_SELECTOR, ".dataTables_info") else ""
                    ) != (pre_info or "")
                )
            except Exception:
                time.sleep(0.8)

            read_state = None
            read_perf = None
            if read_mode in ("observe_only", "state_driven"):
                read_state, read_perf = read_request_list_state(d)
                read_perf["refresh_ms"] = refresh_ms
                if read_mode == "state_driven":
                    st = (read_state.get("state") or "").upper()
                    if st == "EMPTY_CONFIRMED":
                        emit_request_logs(read_state, read_perf, legacy_result="SKIPPED_BY_STATE")
                        return False, "NO_REQUEST_CONFIRMED @ state_driven"
                    if st == "ERROR":
                        emit_request_logs(read_state, read_perf, legacy_result="SKIPPED_BY_STATE")
                        return False, f"REQUEST_READ_ERROR @ state_driven: {read_state.get('exception')}"

            stage = "WAIT_ROWS_BEFORE"

            def wait_rows_and_sample(driver):
                WebDriverWait(driver, 20).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "tbody tr"))
                )
                rows_inner = driver.find_elements(By.CSS_SELECTOR, "tbody tr")
                return rows_inner

            def find_target_row(driver, proj_code_val):
                proj_digits_val = re.sub(r"\D", "", proj_code_val or "")
                samples_inner = []
                target_inner = None
                for attempt_inner in range(3):
                    try:
                        rows_inner = driver.find_elements(By.CSS_SELECTOR, "tbody tr")
                        samples_inner = []
                        for idx, tr in enumerate(rows_inner):
                            tds = tr.find_elements(By.CSS_SELECTOR, "td")
                            full_txt = " ".join([normalize_text(td.text) for td in tds])
                            full_digits = re.sub(r"\D", "", full_txt)
                            if idx < 3:
                                samples_inner.append(full_txt or "(empty)")
                            if (proj_code_val and proj_code_val in full_txt) or (proj_digits_val and proj_digits_val in full_digits):
                                target_inner = tr
                                break
                        if target_inner is not None:
                            break
                    except StaleElementReferenceException:
                        target_inner = None
                        samples_inner = []
                        refocus_first_iframe(driver)
                        time.sleep(0.3)
                return target_inner, samples_inner, proj_digits_val

            rows = wait_rows_and_sample(d)
            if not rows:
                return False, f"ROWS_EMPTY_BEFORE @ {stage}"

            stage = "FIND_TARGET_ROW_BEFORE"
            target, samples, proj_digits = find_target_row(d, proj_code)

            legacy_result = "HAS_REQUEST" if target else "NO_REQUEST"
            if read_mode in ("observe_only", "state_driven") and read_state and read_perf:
                emit_request_logs(read_state, read_perf, legacy_result=legacy_result)

            if not target:
                stage = "RETRY_SEARCH_BEFORE"
                try:
                    pre_info_retry = d.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip()
                except Exception:
                    pre_info_retry = ""
                if refocus_first_iframe(d, reload_url=BUS_NEW_URL):
                    click_search_manual(d)
                try:
                    WebDriverWait(d, 7).until(
                        lambda x: (
                            x.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip()
                            if x.find_elements(By.CSS_SELECTOR, ".dataTables_info") else ""
                        ) != (pre_info_retry or "")
                    )
                except Exception:
                    time.sleep(0.8)
                if not enter_first_iframe(d, reload_url=BUS_NEW_URL):
                    return False, f"IFRAME_REENTER_FAIL @ {stage}"
                rows = wait_rows_and_sample(d)
                if not rows:
                    return False, f"ROWS_EMPTY_BEFORE @ {stage}"
                stage = "FIND_TARGET_ROW_BEFORE"
                target, samples, proj_digits = find_target_row(d, proj_code)

            if not target:
                return False, f"ROW_NOT_FOUND @ {stage}: proj={proj_code}, digits={proj_digits}, sample={samples}"

            stage = "FIND_PROCESS_BUTTON"
            try:
                btn = target.find_element(By.XPATH, ".//button[contains(.,'처리')]")
            except Exception:
                return False, f"BTN_NOT_FOUND @ {stage}"

            stage = "CLICK_PROCESS_BUTTON"
            try:
                d.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                try:
                    btn.click()
                except ElementClickInterceptedException:
                    d.execute_script("arguments[0].click();", btn)
            except Exception as e:
                return False, f"BTN_CLICK_FAIL @ {stage}: {e}"

            stage = "WAIT_CONFIRM_POPUP"

            popup = None
            for _ in range(40):  # 최대 8초 (0.2 * 40)
                try:
                    popup = find_swal_popup(d)
                    if popup:
                        break
                except Exception:
                    popup = None
                time.sleep(0.2)

            if popup is None:
                return False, f"CONFIRM_POPUP_NOT_FOUND @ {stage}"

            stage = "READ_CONFIRM_MESSAGE"
            try:
                bus_msg_title = (popup.find_element(By.CSS_SELECTOR, ".swal2-title").text or "").strip()
            except Exception:
                bus_msg_title = ""
            try:
                bus_msg_body = (popup.find_element(By.CSS_SELECTOR, ".swal2-html-container").text or "").strip()
            except Exception:
                bus_msg_body = ""

            stage = "CLICK_CONFIRM_POPUP"
            try:
                ok_btn = WebDriverWait(popup, 5).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "button.swal2-confirm"))
                )
                try:
                    ok_btn.click()
                except ElementClickInterceptedException:
                    d.execute_script("arguments[0].click();", ok_btn)
            except Exception as e:
                return False, f"CONFIRM_POPUP_CLICK_FAIL @ {stage}: {e}"

            stage = "WAIT_CONFIRM_CLOSE"
            try:
                WebDriverWait(d, 10).until(lambda _:
                    find_swal_popup(d) is None
                )
            except Exception:
                pass

            stage = "RELOAD_IFRAME_AFTER_PROCESS"
            if not enter_first_iframe(d, reload_url=BUS_NEW_URL):
                return False, f"IFRAME_REENTER_FAIL @ {stage}"

            try:
                pre_info2 = d.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip()
            except Exception:
                pre_info2 = ""

            stage = "SET_FILTER_AFTER"
            try:
                set_select_value_and_fire(d, "processYn", "N")
            except Exception:
                pass

            stage = "CLICK_SEARCH_AFTER"
            if not refocus_first_iframe(d, reload_url=BUS_NEW_URL):
                return False, f"IFRAME_LOST_AFTER_SEARCH @ {stage}"
            click_search_manual(d)

            stage = "WAIT_FILTER_APPLY_AFTER"
            try:
                WebDriverWait(d, 7).until(
                    lambda x: (
                        x.find_element(By.CSS_SELECTOR, ".dataTables_info").text.strip()
                        if x.find_elements(By.CSS_SELECTOR, ".dataTables_info") else ""
                    ) != (pre_info2 or "")
                )
            except Exception:
                time.sleep(0.8)

            stage = "WAIT_ROWS_AFTER"
            try:
                WebDriverWait(d, 20).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "tbody tr"))
                )
                rows2 = d.find_elements(By.CSS_SELECTOR, "tbody tr")
            except Exception:
                rows2 = []

            stage = "VERIFY_ROW_REMOVED"
            proj_digits = re.sub(r"\D", "", proj_code or "")
            still_exists = False
            for tr in rows2:
                tds = tr.find_elements(By.CSS_SELECTOR, "td")
                full_txt = " ".join([normalize_text(td.text) for td in tds])
                full_digits = re.sub(r"\D", "", full_txt)
                if (proj_code and proj_code in full_txt) or (proj_digits and proj_digits in full_digits):
                    still_exists = True
                    break

            if still_exists:
                if bus_msg_title or bus_msg_body:
                    return False, f"ROW_STILL_EXISTS_AFTER_PROCESS @ {stage}: title={bus_msg_title}, body={bus_msg_body}"
                return False, f"ROW_STILL_EXISTS_AFTER_PROCESS @ {stage}"

            if bus_msg_title or bus_msg_body:
                return True, f"OK @ {stage}: title={bus_msg_title}, body={bus_msg_body}"
            return True, f"OK @ {stage}"

        except Exception as e:
            return False, f"EXCEPTION @ {stage}: {e}"

        finally:
            if d is not None:
                try:
                    d.quit()
                except Exception:
                    pass

    def _create_group_and_folder(self, proj: str, proj_name: str) -> (bool, str):
        group_name = format_group_name(proj, "Study")
        root_path  = build_root_from_proj(proj)
        study_all  = os.path.join(root_path, "study", "all")

        TEMPLATE_ROOT = r"\\LSK_S010\Study folder\_Template"
        GROUP_OU_PATH = r"OU=Group Project Folder,OU=1.Management Object Group,OU=lskglobal,DC=lskglobal,DC=com"

        ps = []
        ps.append("$ErrorActionPreference='Stop';")
        ps.append(
            f"$g = Get-ADGroup -Filter \"SamAccountName -eq '{psq(group_name)}'\";"
        )
        ps.append(
            f"if (-not $g) {{ "
            f"New-ADGroup -Name '{psq(group_name)}' -SamAccountName '{psq(group_name)}' "
            f"-GroupCategory Security -GroupScope DomainLocal "
            f"-Path '{psq(GROUP_OU_PATH)}' -Description '{psq(proj_name or '')}'; "
            f"Start-Sleep -Seconds 15; "
            f"}} else {{ Write-Host 'GROUP_EXISTS'; }}"
        )
        ps.append(
            f"if (!(Test-Path '{psq(root_path)}')) {{ "
            f"robocopy '{psq(TEMPLATE_ROOT)}' '{psq(root_path)}' *.* /E /COPYALL | Out-Null; }}"
        )
        ps.append(
            f"$paths = @('{psq(root_path)}', '{psq(study_all)}'); "
            "foreach($p in $paths) { "
            "if (Test-Path $p) { "
            "$acl = Get-Acl $p; $unknown = @(); "
            "foreach($ace in $acl.Access) { "
            "$val = $ace.IdentityReference.Value; "
            "if ($val -and $val -match '^S-1-') { $unknown += $val; continue } "
            "try { "
                "$null = $ace.IdentityReference.Translate([System.Security.Principal.NTAccount]); "
            "} catch { "
                "$val = $ace.IdentityReference.Value; "
                "if ($val -and $val -match '^S-1-') { $unknown += $val } "
                "elseif ($val) { $unknown += $val } "
            "} "
            "} "
            "$unknown = $unknown | Sort-Object -Unique; "
            "foreach($sid in $unknown) { "
            "try { icacls \"$p\" /remove \"$sid\" /T /C | Out-Null; } catch {} "
            "} "
            "} "
            "} "
        )
        ps.append(
            f"if (Test-Path '{psq(root_path)}') {{ "
            f"icacls '{psq(root_path)}' /grant '{psq(group_name)}:(ci)(oi)rx' | Out-Null; }}"
        )
        ps.append(
            f"if (Test-Path '{psq(study_all)}') {{ "
            f"icacls '{psq(study_all)}' /grant '{psq(group_name)}:(ci)(oi)rxm' | Out-Null; }}"
        )

        cmd = " ".join(ps)
        ok, out = self._run_pwsh(cmd, timeout_ms=240000)
        return ok, out

    def run(self):
        total = len(self.items)
        ok_cnt = fail_cnt = 0

        if not self.creds.get("id") or not self.creds.get("pw"):
            self.error.emit("BUS 계정(아이디/비밀번호) 누락")
            self.finished.emit(0, total)
            return

        for idx, it in enumerate(self.items, start=1):
            if self._stop:
                self.finished.emit(ok_cnt, fail_cnt)
                return

            proj = it.get("proj","").strip()
            name = it.get("name","").strip()

            self.progress.emit(idx-1, total, f"[{idx}/{total}] {proj} 폴더/권한 생성 중…")
            ok, msg = self._create_group_and_folder(proj, name)
            if not ok:
                fail_cnt += 1
                self.progress.emit(idx, total, f"[{idx}/{total}] {proj} 실패: {msg}")
                continue

            self.progress.emit(idx-1, total, f"[{idx}/{total}] {proj} BUS 처리 중…")
            b_ok, b_msg = self._bus_click_process(proj)
            if b_ok:
                ok_cnt += 1
                self.progress.emit(idx, total, f"[{idx}/{total}] {proj} 완료")
            else:
                fail_cnt += 1
                self.progress.emit(idx, total, f"[{idx}/{total}] {proj} BUS 실패: {b_msg}")

        self.finished.emit(ok_cnt, fail_cnt)

class BadgeToolButton(QWidget):
    def __init__(self, *args, **kwargs):
        target = None
        parent = None
        if len(args) == 1:
            parent = args[0]
        elif len(args) >= 2:
            if isinstance(args[0], QAbstractButton):
                target = args[0]
                parent = args[1]
            else:
                parent = args[1]
        else:
            parent = kwargs.get("parent", None)

        super().__init__(parent)
        self._target = None
        self._offset = QPoint(0, 1)
        self._value = 0
        self._lbl = QLabel(self)
        self._lbl.setObjectName("notifBadge")
        self._lbl.setAlignment(Qt.AlignCenter)
        f = self._lbl.font()
        f.setPointSize(7)
        f.setBold(True)
        self._lbl.setFont(f)
        self._base_px = 7
        self._apply_style(self._base_px)
        self._lbl.setVisible(False)

        self.setAttribute(Qt.WA_TransparentForMouseEvents, True)
        self._lbl.setAttribute(Qt.WA_TransparentForMouseEvents, True)

        if target is not None:
            self.setTarget(target)

    def setTarget(self, btn: QAbstractButton):
        if self._target is not None:
            try:
                self._target.removeEventFilter(self)
            except Exception:
                pass
        self._target = btn
        self._lbl.setParent(self._target)
        self._lbl.raise_()
        if self._target is not None:
            self._target.installEventFilter(self)
        self._reposition()

    def setBadge(self, n: int):
        self._value = int(n or 0)
        if self._value <= 0:
            self._lbl.hide()
            return
        self._lbl.setText("99+" if self._value >= 100 else str(self._value))
        self._lbl.adjustSize()
        self._lbl.show()
        self._lbl.raise_()
        self._reposition()

    def setOffset(self, dx: int, dy: int):
        self._offset = QPoint(dx, dy)
        self._reposition()

    def setBadgeSize(self, px: int = 1):
        self._base_px = px
        self._apply_style(px)
        self._lbl.adjustSize()
        self._reposition()

    def eventFilter(self, obj, ev):
        if obj is self._target and ev.type() in (ev.Resize, ev.Move, ev.Show):
            self._reposition()
        return super().eventFilter(obj, ev)

    def _reposition(self):
        if not (self._target and self._lbl.isVisible()):
            return
        bw = max(self._base_px, self._lbl.width())
        x = self._target.width() - bw + self._offset.x()
        y = self._offset.y()
        self._lbl.move(x, y)

    def _apply_style(self, px: int):
        r = px // 2
        pad = max(1, px // 4)
        maxw = px * 3

        self._lbl.setStyleSheet(f"""
            QLabel#notifBadge {{
                color: white;
                background: #E23;
                border-radius: {r}px;
                min-width: {px}px;
                min-height: {px}px;
                max-width: {maxw}px;
                padding-left: {pad}px;
                padding-right: {pad}px;
            }}
        """)

class AccessManager(QMainWindow):
    trigger_session_start = pyqtSignal()
    trigger_session_download = pyqtSignal()
    trigger_session_cancel = pyqtSignal()
    trigger_session_stop = pyqtSignal()
    trigger_session_process = pyqtSignal(object)
    trigger_watcher_start = pyqtSignal()
    trigger_watcher_stop  = pyqtSignal()
    trigger_watcher_collect = pyqtSignal()
    COL_SELECT  = 0
    COL_KIND    = 1
    COL_REQTYPE = 2
    COL_USER    = 3
    COL_NAME    = 4
    COL_PROJ    = 5
    COL_LV2     = 6
    COL_LV3     = 7
    COL_DEPT    = 8
    COL_ROLE    = 9
    COL_STATUS  = 10

    EDITABLE_COLS = {COL_KIND, COL_REQTYPE, COL_USER, COL_NAME, COL_PROJ, COL_LV2, COL_LV3, COL_DEPT, COL_ROLE}

    def _show_help(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("기능 설명")
        edit = QTextEdit(dlg)
        edit.setReadOnly(True)
        edit.setPlainText(HELP_TEXT)
        btns = QDialogButtonBox(QDialogButtonBox.Close, parent=dlg)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        lay = QVBoxLayout(dlg)
        lay.addWidget(edit)
        lay.addWidget(btns)
        dlg.resize(720, 560)
        dlg.exec_()

    def apply_theme(self, mode: str):
        theme = THEMES.get(mode, THEMES["light"])
        text_color = "#111111" if mode == "light" else "#E0E0E0"
        disabled_bg = "rgba(176,190,197,0.45)" if mode == "light" else "rgba(66,66,66,0.50)"
        disabled_tx = "#9aa3a8"
        disabled_border = "#c7c7c7" if mode == "light" else "#5a5a5a"

        qss = f"""
        QWidget {{
            background-color: {theme['bg']};
            color: {text_color};
        }}

        QPushButton {{
            background-color: {theme['btn']};
            border: 1px solid #9e9e9e;
            border-radius: 6px;
            padding: 6px 12px;
            color: {text_color};
        }}
        QPushButton:hover {{
            background-color: {theme['hover']};
        }}
        QPushButton:pressed {{
            background-color: {theme['press']};
        }}

        QPushButton:disabled,
        QPushButton:disabled:hover,
        QPushButton:disabled:pressed {{
            background-color: rgba(176,190,197,0.45);
            color: #9aa3a8;
            border: 1px solid {theme['panel_border']};
        }}

        QToolButton:disabled {{
            color: #9aa3a8;
        }}

        QLineEdit:disabled,
        QComboBox:disabled,
        QCheckBox:disabled,
        QTextEdit:disabled {{
            color: #9aa3a8;
            background-color: transparent;
        }}

        QHeaderView::section {{
            background-color: {theme['btn']};
            color: {text_color};
            border: 0px;
            padding: 4px 6px;
        }}

        QHeaderView::section:vertical {{
            background-color: {theme['btn']};
            color: {text_color};
            border: 0px;
            padding: 0px 6px;
            border-radius: 0px;
        }}

        QTableCornerButton::section {{
            background-color: {theme['btn']};
            border: 0px;
            padding: 0px;
            border-top-left-radius: 8px;
        }}

        QTableView, QTableWidget, QTextEdit {{
            background: {theme['panel']};
            border: 1px solid {theme['panel_border']};
            border-radius: 8px;
            selection-background-color: {theme['select']};
            selection-color: {text_color};
        }}

        QTableView, QTableWidget {{
            gridline-color: {theme['panel_border']};
            alternate-background-color: {theme['alt']};
        }}

        QTableWidget::item, QTableView::item {{
            padding: 4px 6px;
        }}

        QAbstractScrollArea {{
            background: transparent;
        }}

        QMenu {{
            background-color: {theme['bg']};
            color: {text_color};
            border: 1px solid #9e9e9e;
            border-radius: 6px;
            padding: 4px;
        }}
        QMenu::separator {{
            height: 1px;
            background: rgba(0,0,0,0.15);
            margin: 4px 8px;
        }}
        QMenu::item {{
            padding: 6px 12px;
            border-radius: 4px;
            background: transparent;
        }}
        QMenu::item:selected {{
            background: {theme['hover']};
            color: {text_color};
        }}

        QComboBox QAbstractItemView {{
            background-color: {theme['bg']};
            color: {text_color};
            selection-background-color: {theme['hover']};
            selection-color: {text_color};
            outline: 0;
        }}

        QHeaderView::section:horizontal:last {{
            border-top-right-radius: 8px;
        }}

        QHeaderView::section:horizontal:first {{
            border-top-left-radius: 0px;
        }}
        """
        self.setStyleSheet(qss)
        self.current_theme = mode
        self._refresh_theme_button_emoji()

        if hasattr(self, "btn_theme") and self.btn_theme:
            self.btn_theme.setText("🌙" if mode == "light" else "🌞")
            self.btn_theme.setToolTip("다크 모드로 전환" if mode == "light" else "라이트 모드로 전환")

        if hasattr(self, "table") and self.table and hasattr(self.table.horizontalHeader(), "_reposition"):
            self.table.horizontalHeader()._reposition()
        
    def _toggle_theme(self):
        next_mode = "dark" if getattr(self, "current_theme", "light") == "light" else "light"
        self.apply_theme(next_mode)
        try:
            data = {}
            if os.path.exists(CONF_FILE):
                with open(CONF_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
            data["theme"] = next_mode
            with open(CONF_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def __init__(self):
        super().__init__()
        self._notif_counts = {"신규미완료": 0, "진행-부여": 0, "진행-제거": 0, "종료-부여": 0, "종료-제거": 0}
        self.setWindowTitle(f"{APP_NAME}")
        self.resize(900, 600)
        self._init_ui()
        self._notify_refresh_mins = 10
        try:
            if os.path.exists(CONF_FILE):
                with open(CONF_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self._notify_refresh_mins = int(data.get("notify_refresh_min", self._notify_refresh_mins))
        except Exception:
            pass
        self.watch_thread = QThread(self)
        self.watch_session = BusWatcher(DL_DIR)
        self.watch_session.moveToThread(self.watch_thread)
        self.watch_thread.start()
        self.trigger_watcher_start.connect(self.watch_session.start, type=Qt.QueuedConnection)
        self.trigger_watcher_stop.connect(self.watch_session.stop, type=Qt.QueuedConnection)
        self.trigger_watcher_collect.connect(self.watch_session.collect_counts, type=Qt.QueuedConnection)
        self.watch_session.countsReady.connect(self._on_counts_ready)
        self.notify_timer = QTimer(self)
        self.notify_timer.setInterval(max(5, int(self._notify_refresh_mins)) * 60 * 1000)
        self.notify_timer.timeout.connect(self.refresh_notifications)
        self.notify_timer.start()
        self._ensure_logo_and_set_title()
        self.session_thread = QThread(self)
        self.session = BusSessionManager(DL_DIR)
        self.session.moveToThread(self.session_thread)
        self.session.readyChanged.connect(self._on_session_ready)
        self.session.downloaded.connect(self._on_session_downloaded)
        self.session.processed.connect(self._on_session_processed)
        self.session.busyChanged.connect(self._on_session_busy)
        self.session_thread.start()
        self.watch_session.readyChanged.connect(self._on_watcher_ready)
        self.trigger_session_start.connect(self.session.start, type=Qt.QueuedConnection)
        self.trigger_session_download.connect(self.session.download_list, type=Qt.QueuedConnection)
        self.trigger_session_process.connect(self.session.process, type=Qt.QueuedConnection)
        self.trigger_session_cancel.connect(self.session.cancel_current, type=Qt.QueuedConnection)
        self.trigger_session_stop.connect(self.session.stop, type=Qt.QueuedConnection)
        QTimer.singleShot(1000, self.refresh_notifications)
        self.ps_path, self.ps_kind = self._pick_powershell()
        self.proc = QProcess(self)
        self.proc.readyReadStandardOutput.connect(self._ps_ready_out)
        self.proc.readyReadStandardError.connect(self._ps_ready_err)
        self.proc.finished.connect(self._ps_finished)
        self.run_queue = []
        self.total_jobs = 0
        self.done_jobs = 0
        self.current_row = -1
        self.current_seq = 0
        self.current_mode = ""
        self.setAcceptDrops(True)
        self._load_creds()
        if self.creds.get("id") and self.creds.get("pw"):
            self.watch_session.set_creds(self.creds["id"], self.creds["pw"])
            self.trigger_watcher_start.emit()
        self._log(f"{APP_NAME} 실행")
        self._buf_out = {}
        self.remove_fail_tolerance = int(self.creds.get("remove_fail_tol", 5))
        self.auto_complete_after_add = True
        self._waiting_for_bus = False
        self._pending_after_add_row = None
        self.bus_seq = 0
        self._ignore_bus_results = False
        self.stop_requested = False
        self._auto_create_group_decided = None
        self._retrying_after_group_create = False
        self._bus_mode = False
        self._bus_queue = []
        self._bus_done = 0
        self._bus_total = 0

        saved_theme = self._read_saved_theme()
        self.current_theme = saved_theme or "light"
        self.apply_theme(self.current_theme)

        self.debug_enabled = bool(self.creds.get("debug", False))
        self.session.set_debug(self.debug_enabled, DEBUG_DIR)
        
    def _init_ui(self):
        main_layout = QVBoxLayout()
        header_bar = QHBoxLayout()
        self.header_bar = header_bar
        header_bar.setContentsMargins(0, 0, 0, 0)
        header_bar.setSpacing(6)

        header_bar.addStretch()
        self._title_widget = QLabel("")
        header_bar.addWidget(self._title_widget, 0, Qt.AlignVCenter)
        header_bar.addStretch()

        self.btn_notif = QToolButton(self)
        self.btn_notif.setText("💡")
        self.btn_notif.setToolTip("요청건 알림")
        self.btn_notif.setAutoRaise(True)
        self.btn_notif.setStyleSheet("""
            QToolButton {
                border: none;
                background: transparent;
                font-size: 15px;
                padding: 0px;
            }
            QToolButton:hover {
                background: transparent;    /* 마우스 올렸을 때 배경 투명 */
            }
            QToolButton:pressed {
                background: transparent;    /* 클릭 중에도 투명 */
            }
            QToolButton:checked {
                background: transparent;    /* 체크 상태도 투명 */
            }
            QToolButton:focus {
                outline: none;              /* 포커스 테두리 제거 */
            }
        """)

        self.btn_notif.clicked.connect(self._open_notif_popup)
        self.badge_notif = BadgeToolButton(self.btn_notif, self)
        self.badge_notif.setBadgeSize(9)
        self.badge_notif.setOffset(0,0)

        self.btn_theme = QToolButton(self)
        self.btn_theme.setAutoRaise(True)
        self.btn_theme.setCursor(Qt.PointingHandCursor)
        self.btn_theme.setToolTip("테마 전환")
        self.btn_theme.setStyleSheet("""
            QToolButton {
                border: none;
                background: transparent;
                font-size: 15px;
                padding: 0 6px;
            }
            QToolButton:hover {
                background: transparent;
            }
        """)
        header_bar.addWidget(self.btn_theme, 0, Qt.AlignRight | Qt.AlignVCenter)
        header_bar.addWidget(self.btn_notif, 0, Qt.AlignRight | Qt.AlignVCenter)
        header_w = QWidget()
        header_w.setLayout(header_bar)
        main_layout.addWidget(header_w)

        file_bar = QHBoxLayout()
        #self.file_label = QLabel("요청 확인 버튼 클릭 또는 엑셀 파일 불러오기, 수동으로 리스트 입력")
        #self.file_label.setFont(QFont("Segoe UI", 9))

        self.btn_newcheck = QPushButton("𝙉 신규 확인")
        self.btn_newcheck.setFont(QFont("Segoe UI", 9))
        self.btn_newcheck.clicked.connect(self.open_new_viewer) 

        self.btn_request = QPushButton("🔍 요청 확인")
        self.btn_request.setFont(QFont("Segoe UI", 9))
        self.btn_request.clicked.connect(self._request_and_import)

        self.btn_manual = QPushButton("📝 수동 입력")
        self.btn_manual.setFont(QFont("Segoe UI", 9))
        self.btn_manual.clicked.connect(self.open_manual_dialog)
        
        self.btn_file = QPushButton("📂 파일 선택")
        self.btn_file.setFont(QFont("Segoe UI", 9))
        self.btn_file.clicked.connect(self.choose_file)

        self.btn_settings = QPushButton("⚙ 설정")
        self.btn_settings.setFont(QFont("Segoe UI", 9))
        self.btn_settings.clicked.connect(self._open_settings)
        
        #file_bar.addWidget(self.file_label)
        file_bar.addStretch()
        file_bar.addWidget(self.btn_newcheck)
        file_bar.addWidget(self.btn_request)
        file_bar.addWidget(self.btn_manual)
        file_bar.addWidget(self.btn_file)
        file_bar.addWidget(self.btn_settings)
        main_layout.addLayout(file_bar)

        self.table = CopyTable()
        self.table.setProperty("hasRows", False)
        header = CheckBoxHeader(Qt.Horizontal, self.table)
        self.table.setHorizontalHeader(header)
        header.stateChanged.connect(self.toggle_all_rows)

        self.table.setFont(QFont("Consolas", 9))
        self.table.setColumnCount(11)
        self.table.setHorizontalHeaderLabels([
            "", "구분", "요청사항", "사번", "이름", "프로젝트코드","Level2","Level3","부서","Role","상태"
        ])
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked | QAbstractItemView.EditKeyPressed)
        self.table.cellChanged.connect(self._on_cell_changed)

        hv = self.table.horizontalHeader()
        hv.setSectionResizeMode(QHeaderView.ResizeToContents)
        hv.setStretchLastSection(True)

        hv.setSectionResizeMode(self.COL_SELECT, QHeaderView.Fixed)
        self.table.setColumnWidth(self.COL_SELECT, 28)
        hv.setMinimumSectionSize(20)

        vh = self.table.verticalHeader()
        vh.setSectionResizeMode(QHeaderView.Fixed)
        vh.setDefaultSectionSize(28)
        vh.setMinimumSectionSize(22)
        vh.setFixedWidth(28)
        self.table.setCornerButtonEnabled(True)

        header.sectionResized.connect(lambda *_: header.updateSection(self.COL_SELECT))
        try:
            header.geometriesChanged.connect(lambda: header.updateSection(self.COL_SELECT))
        except Exception:
            pass
        
        main_layout.addWidget(self.table)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._open_table_menu)
        
        btn_bar = QHBoxLayout()

        self.btn_clear_log = QPushButton("🗑 로그 지우기")
        self.btn_clear_log.setFont(QFont("Segoe UI", 9))
        self.btn_clear_log.setToolTip("아래 로그 창 내용을 비웁니다")
        self.btn_clear_log.clicked.connect(self._clear_log)

        self.chk_dry = QCheckBox("Dry Run")
        self.chk_dry.setFont(QFont("Segoe UI", 9))

        self.chk_auto_complete = QCheckBox("실행 후 자동 완료처리")
        self.chk_auto_complete.setFont(QFont("Segoe UI", 9))
        self.chk_auto_complete.setChecked(True)
        self.chk_auto_complete.stateChanged.connect(
            lambda _: setattr(self, "auto_complete_after_add", self.chk_auto_complete.isChecked())
        )

        self.btn_run_execute = QPushButton("> 실행")
        self.btn_run_execute.setFont(QFont("Segoe UI", 9))
        self.btn_run_execute.clicked.connect(self.run_execute)

        self.btn_run_complete = QPushButton("v 완료 처리")
        self.btn_run_complete.setFont(QFont("Segoe UI", 9))
        self.btn_run_complete.clicked.connect(self.run_complete)

        self.btn_stop = QPushButton("x 중지")
        self.btn_stop.setFont(QFont("Segoe UI", 9))
        self.btn_stop.setToolTip("현재 실행 중인 작업과 대기 중인 모든 작업을 중지합니다")
        self.btn_stop.setEnabled(False)
        self.btn_stop.clicked.connect(self._stop_all)

        self.btn_theme.clicked.connect(self._toggle_theme)

        btn_bar.addWidget(self.btn_clear_log)
        btn_bar.addStretch()
        btn_bar.addWidget(self.chk_dry)
        btn_bar.addWidget(self.chk_auto_complete)
        btn_bar.addWidget(self.btn_run_execute)
        btn_bar.addWidget(self.btn_run_complete)
        btn_bar.addWidget(self.btn_stop)
        main_layout.addLayout(btn_bar)
        
        self.log = QTextEdit()
        self.log.setAcceptRichText(False)
        self.log.setLineWrapMode(QTextEdit.WidgetWidth)
        self.log.setWordWrapMode(QTextOption.WrapAnywhere)
        self.log.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.log.setFont(QFont("Consolas", 9))
        self.log.setReadOnly(True)
        main_layout.addWidget(self.log)

        container = QWidget()
        container.setLayout(main_layout)
        self.setCentralWidget(container)

        sb = super().statusBar()
        if sb is None:
            sb = QStatusBar(self)
            self.setStatusBar(sb)
        self._statusbar = sb

        self.prg = QProgressBar()
        self.prg.setRange(0,0)
        self.prg.setFixedWidth(160)
        self.prg.hide()
        self.status_label = QLabel("")
        self._statusbar.addPermanentWidget(self.status_label)
        self._statusbar.addPermanentWidget(self.prg)

    def refresh_notifications(self):
        if hasattr(self, "watch_session") and self.watch_session and self.watch_session.is_ready():
            self.trigger_watcher_collect.emit()
            return
        self._on_counts_ready({"신규미완료":0,"진행-부여":0,"진행-제거":0,"종료-부여":0,"종료-제거":0})

    @pyqtSlot(dict)
    def _on_counts_ready(self, counts: dict):
        counts = counts or {}
        def _toi(x):
            try:
                return int(str(x).strip())
            except Exception:
                return 0

        self._notif_counts = {
            "신규미완료": _toi(counts.get("신규미완료", 0)),
            "진행-부여": _toi(counts.get("진행-부여", 0)),
            "진행-제거": _toi(counts.get("진행-제거", 0)),
            "종료-부여": _toi(counts.get("종료-부여", 0)),
            "종료-제거": _toi(counts.get("종료-제거", 0)),
        }

        total = sum(self._notif_counts.values())

        QTimer.singleShot(0, lambda t=total: self.badge_notif.setBadge(t))

    def _on_watcher_ready(self, ok: bool, msg: str):
        if hasattr(self, "btn_notif"):
            self.btn_notif.setEnabled(bool(ok))
            self.btn_notif.setToolTip("요청건 알림" if ok else "세션 준비 안됨")

        if ok:
            QTimer.singleShot(0, lambda: self.trigger_watcher_collect.emit())
        else:
            self._on_counts_ready({"신규미완료":0,"진행-부여":0,"진행-제거":0,"종료-부여":0,"종료-제거":0})

    def _open_notif_popup(self):
        if not hasattr(self, "_notif_counts") or not isinstance(self._notif_counts, dict):
            self._notif_counts = {"신규미완료":0,"진행-부여":0,"진행-제거":0,"종료-부여":0,"종료-제거":0}
        c = self._notif_counts
        menu = QMenu(self)
        n_new  = int(self._notif_counts.get("신규미완료", 0))
        n_pg_g = int(self._notif_counts.get("진행-부여", 0))
        n_pg_r = int(self._notif_counts.get("진행-제거", 0))
        n_ed_g = int(self._notif_counts.get("종료-부여", 0))
        n_ed_r = int(self._notif_counts.get("종료-제거", 0))
        menu.addAction(f"폴더 생성 요청: {n_new}")
        menu.addAction(f"진행-부여: {n_pg_g}")
        menu.addAction(f"진행-제거: {n_pg_r}")
        menu.addAction(f"종료-부여: {n_ed_g}")
        menu.addAction(f"종료-제거: {n_ed_r}")
        pos = self.btn_notif.mapToGlobal(self.btn_notif.rect().bottomRight())
        menu.popup(pos)

    def open_new_viewer(self):
        def _after_ready(ok: bool):
            if not ok:
                return
            self._newdlg = NewItemsViewer(self)
            self._newdlg.btn_refresh.clicked.connect(self._load_new_items)
            self._newdlg.show()
            self._load_new_items()

        self.ensure_bus_session_async("폴더 생성 요청 조회", _after_ready)

    def _load_new_items(self):
        try:
            self.session.newDownloaded.disconnect(self._on_new_downloaded)
        except Exception:
            pass
        self.session.newDownloaded.connect(self._on_new_downloaded)
        self.status_label.setText("폴더 생성 요청 조회 중")
        self._set_running_ui(True)
        self.trigger_session_start.emit()
        self.session.download_new_list()

    def _on_new_downloaded(self, path: str, err: str):
        self._set_running_ui(False)
        if err:
            QMessageBox.critical(self, "오류", f"폴더 생성 요청 로드 실패: {err}")
            return
        header, data = _parse_html_best_table(path)
        if not header:
            QMessageBox.warning(self, "안내", "표 데이터를 찾지 못했습니다.")
            return
        if getattr(self, "_newdlg", None):
            self._newdlg.set_data(header, data)
        self.status_label.setText(f"폴더 생성 요청 {len(data)}건")

    def _create_group_and_base_acl(self, proj: str, lv2: str, lv3: str) -> (bool, str):
        try:
            group_name = format_group_name(proj, lv2)
            is_isolated = (normalize_lv2(lv2) == "Isolated")

            base_group = format_group_name(proj, "Study")

            root_path = build_root_from_proj(proj)
            isolated_path = os.path.join(root_path, "Isolated")

            lines = []
            lines.append("$ErrorActionPreference='Stop';")

            if is_isolated:
                lines.append(f"$desc = try {{ (Get-ADGroup -Identity '{psq(base_group)}' -Properties Description).Description }} catch {{ '' }};")
                lines.append(
                    "New-ADGroup "
                    f"-Name '{psq(group_name)}' -SamAccountName '{psq(group_name)}' "
                    "-GroupCategory Security -GroupScope DomainLocal "
                    f"-Path '{psq(GROUP_OU_PATH)}' -Description $desc;"
                )
                lines.append("Start-Sleep -s 15;")
                lines.append(f"if (Test-Path '{psq(root_path)}') {{ icacls '{psq(root_path)}' /grant '{psq(group_name)}:rx' | Out-Null; }}")
                lines.append(f"if (Test-Path '{psq(isolated_path)}') {{ icacls '{psq(isolated_path)}' /grant '{psq(group_name)}:rx' | Out-Null; }}")
            else:
                lines.append(
                    "New-ADGroup "
                    f"-Name '{psq(group_name)}' -SamAccountName '{psq(group_name)}' "
                    "-GroupCategory Security -GroupScope DomainLocal "
                    f"-Path '{psq(GROUP_OU_PATH)}' -Description '';"
                )
                lines.append(f"if (!(Test-Path '{psq(root_path)}')) {{ robocopy '{psq(TEMPLATE_ROOT)}' '{psq(root_path)}' *.* /E /COPYALL | Out-Null; }}")
                lines.append(f"if (Test-Path '{psq(root_path)}') {{ icacls '{psq(root_path)}' /grant '{psq(group_name)}:(ci)(oi)rx' | Out-Null; }}")
                lines.append(f"if (Test-Path '{psq(os.path.join(root_path,'study','all'))}') {{ icacls '{psq(os.path.join(root_path,'study','all'))}' /grant '{psq(group_name)}:(ci)(oi)rxm' | Out-Null; }}")

            cmd = " ".join(lines)
            wrapped = self._wrap_cmd_utf8(cmd)
            p = QProcess(self)
            args = ["-NoLogo","-NoProfile","-ExecutionPolicy","Bypass","-Command", wrapped]
            p.start(self.ps_path, args)

            if not p.waitForStarted(30000):
                return False, "PowerShell 시작 실패"

            if not p.waitForFinished(120000):
                try:
                    p.kill()
                    p.waitForFinished(3000)
                except Exception:
                    pass
                return False, "PowerShell 실행 시간 초과"

            rc = p.exitCode()
            out = bytes(p.readAllStandardOutput()).decode("utf-8","ignore")
            err = bytes(p.readAllStandardError()).decode("utf-8","ignore")
            if rc != 0:
                return False, (err.strip() or out.strip() or f"New-ADGroup 실패(code={rc})")

            self._log(f"[보안그룹 생성] {group_name} 완료")
            return True, "OK"
        except Exception as e:
            return False, f"예외: {e}"

    def _refresh_has_rows(self):
        has_rows = self.table.rowCount() > 0

        self.table.setProperty("hasRows", has_rows)
        hh = self.table.horizontalHeader()
        vh = self.table.verticalHeader()
        for w in (self.table, hh, vh):
            try:
                w.setProperty("hasRows", has_rows)
                st = w.style()
                st.unpolish(w); st.polish(w)
                w.update()
            except Exception:
                pass

    def _refresh_theme_button_emoji(self):
        if hasattr(self, "btn_theme") and self.btn_theme:
            self.btn_theme.setText("🌙" if self.current_theme == "light" else "🌞")
            self.btn_theme.setToolTip("다크 모드로 전환" if self.current_theme == "light" else "라이트 모드로 전환")

    def _read_saved_theme(self) -> str:
        try:
            if os.path.exists(CONF_FILE):
                with open(CONF_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    t = (data.get("theme") or "").strip().lower()
                    if t in ("light", "dark"):
                        return t
        except Exception:
            pass
        return "light"

    def _ensure_logo_and_set_title(self):
        try:
            os.makedirs(CONF_DIR, exist_ok=True)

            target_logo = os.path.join(CONF_DIR, LOGO_FILE)
            if not os.path.exists(target_logo):
                bundled_logo = resource_path(os.path.join("assets", LOGO_FILE))
                if os.path.exists(bundled_logo):
                    shutil.copyfile(bundled_logo, target_logo)

            target_icon = os.path.join(CONF_DIR, ICON_FILE)
            if not os.path.exists(target_icon):
                bundled_icon = resource_path(os.path.join("assets", ICON_FILE))
                if os.path.exists(bundled_icon):
                    shutil.copyfile(bundled_icon, target_icon)

            self._apply_title_logo(target_logo if os.path.exists(target_logo) else None)

            if os.path.exists(target_icon):
                app = QApplication.instance()
                icon = QIcon(target_icon)
                self.setWindowIcon(icon)
                if app:
                    app.setWindowIcon(icon)

        except Exception as e:
            self._apply_title_logo(None)
            self._log(f"[로고/아이콘 준비 오류] {e}")

    def _apply_title_logo(self, logo_path: str | None):
        try:
            if hasattr(self, "_title_widget") and self._title_widget is not None:
                self._title_widget.setParent(None)
        except Exception:
            pass

        if logo_path and os.path.exists(logo_path):
            lbl = QLabel()
            pm = QPixmap(logo_path)
            if not pm.isNull():
                pm = pm.scaledToHeight(48, Qt.SmoothTransformation)
                lbl.setPixmap(pm)
            else:
                lbl.setText(APP_NAME)
                lbl.setFont(QFont("Segoe UI", 18, QFont.Bold))
            lbl.setCursor(Qt.PointingHandCursor)
            lbl.mousePressEvent = lambda e: self._show_help()
            widget = lbl
        else:
            txt = QLabel(APP_NAME)
            txt.setFont(QFont("Segoe UI", 18, QFont.Bold))
            txt.setAlignment(Qt.AlignCenter)
            txt.setCursor(Qt.PointingHandCursor)
            txt.mousePressEvent = lambda e: self._show_help()
            widget = txt

        self.header_bar.insertWidget(1, widget, 0, Qt.AlignVCenter)
        self._title_widget = widget

    def _ensure_bus_session(self, purpose: str = "완료 처리") -> bool:
        if not self.creds.get("id") or not self.creds.get("pw"):
            self._open_settings()
            if not self.creds.get("id") or not self.creds.get("pw"):
                QMessageBox.warning(self, "알림", "아이디/비밀번호가 필요합니다.")
                return False

        self.session.set_creds(self.creds["id"], self.creds["pw"])

        if self.session.is_ready():
            return True

        self._set_running_ui(True)
        self.status_label.setText(f"BUS 세션 준비 중 ({purpose})")
        self.trigger_session_start.emit()

        end = time.time() + 60
        while time.time() < end:
            QApplication.processEvents()
            if self.session.is_ready():
                self.status_label.setText("세션 준비됨")
                return True
            time.sleep(0.05)

        self._set_running_ui(False)
        QMessageBox.critical(self, "오류", "세션 초기화 실패")
        return False

    def ensure_bus_session_async(self, purpose: str, on_ready):
        if not self.creds.get("id") or not self.creds.get("pw"):
            self._open_settings()
            if not self.creds.get("id") or not self.creds.get("pw"):
                QMessageBox.warning(self, "알림", "아이디/비밀번호가 필요합니다.")
                if on_ready:
                    try:
                        on_ready(False)
                    except Exception:
                        pass
                return

        self.session.set_creds(self.creds["id"], self.creds["pw"])

        if self.session.is_ready():
            if on_ready:
                try:
                    on_ready(True)
                except Exception:
                    pass
            return

        if not hasattr(self, "_pending_bus_callbacks"):
            self._pending_bus_callbacks = []
        if on_ready:
            self._pending_bus_callbacks.append(on_ready)

        self._set_running_ui(True)
        self.status_label.setText(f"BUS 세션 준비 중 ({purpose})")
        self.trigger_session_start.emit()

    def _start_next_bus_item(self):
        if self.stop_requested or not self._bus_queue:
            self._bus_mode = False
            self._set_running_ui(False)
            self.status_label.setText("완료" if not self.stop_requested else "중지됨")
            self.stop_requested = False
            return

        t = self._bus_queue.pop(0)
        row = t.get('row', -1)
        if 0 <= row < self.table.rowCount():
            it = self.table.item(row, self.COL_STATUS)
            if it:
                it.setText("완료 처리중")

        self.trigger_session_process.emit([t])

    def run_complete(self):
        self._ignore_bus_results = False
        self._waiting_for_bus = False
        self._pending_after_add_row = None
        self.stop_requested = False
        self.auto_complete_after_add = self.chk_auto_complete.isChecked()
        
        if self.table.rowCount() == 0:
            QMessageBox.information(self, "알림", "대상이 없습니다.")
            return

        if self.session.is_busy():
            QMessageBox.warning(self, "알림", "다른 작업이 실행 중입니다. 잠시 후 다시 시도하세요.")
            return

        if not self._ensure_bus_session("완료 처리"):
            return

        allowed_status = {"추가완료", "제거완료", "DryRun"}
        targets = []
        skipped = []

        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, self.COL_SELECT)
            cb = w.findChild(QCheckBox) if w else None
            if not (cb and cb.isChecked()):
                continue

            st = (self.table.item(r, self.COL_STATUS).text().strip()
                  if self.table.item(r, self.COL_STATUS) else "")

            if st not in allowed_status:
                skipped.append(f"{r+1}행: 상태가 '{st}'이므로 완료처리 불가")
                continue

            kind = (self._get(r, self.COL_KIND) or "").strip()

            user = self._get(r, self.COL_USER)
            proj = self._get(r, self.COL_PROJ)
            lv2  = self._get(r, self.COL_LV2)
            lv3  = self._get(r, self.COL_LV3)
            path = "" if kind == "종료" else build_path_l3(proj, lv2, lv3)

            if kind == "종료":
                if not (user and proj):
                    skipped.append(f"{r+1}행: (종료) 필수 정보 누락(user/proj)")
                    continue
                path = build_closed_path_from_proj(proj)
            else:
                if not (user and proj and lv2 and lv3):
                    skipped.append(f"{r+1}행: (진행) 필수 정보 누락(user/proj/lv2/lv3)")
                    continue

            reqtype = (self._get(r, self.COL_REQTYPE) or REQ_GRANT).strip()
            targets.append({
                'row': r, 'kind': kind, 'user': user, 'proj': proj, 'lv2': lv2, 'lv3': lv3, 'path': path, 'req': reqtype
            })

        if skipped:
            QMessageBox.information(self, "완료 제외 안내", "\n".join(skipped))

        if not targets:
            return

        deduped = []
        seen = set()
        for t in targets:
            key = (t['user'], t['proj'], t['lv2'], t['lv3'], (t.get('path') or '').split('\\')[-1])
            if key in seen:
                continue
            seen.add(key)
            deduped.append(t)
        targets = deduped

        if not targets:
            self._set_running_ui(False)
            self.status_label.setText("완료 처리할 항목 없음")
            return

        self.total_jobs = 0
        self.done_jobs = 0

        self._set_running_ui(True)
        self.status_label.setText(f"완료 처리 중 (대상 {len(targets)}건)")
        self._bus_mode = True
        self._bus_queue = targets[:]
        self._bus_total = len(self._bus_queue)
        self._bus_done = 0
        self._set_running_ui(True)
        self.status_label.setText(f"완료 처리 중 {self._bus_done}/{self._bus_total}")
        self._start_next_bus_item()

    def _on_session_processed(self, results):
        if getattr(self, "_ignore_bus_results", False):
            return
        is_auto_wait = getattr(self, "_waiting_for_bus", False)
        pending_row = getattr(self, "_pending_after_add_row", None)

        if (not is_auto_wait) and (not self.run_queue) and (self.proc.state() != QProcess.Running):
            self._set_running_ui(False)

        ok_cnt = fail_cnt = 0
        lines = []
        for r in (results or []):
            row = r.get('row', -1)
            ok = bool(r.get('ok'))
            msg = (r.get('msg') or "").strip()

            if 0 <= row < self.table.rowCount():
                cur = self.table.item(row, self.COL_STATUS)
                if cur and cur.text().strip() == "중지됨":
                    continue

            if 0 <= row < self.table.rowCount():
                it = self.table.item(row, self.COL_STATUS)
                if it:
                    if ok:
                        req = (self._get(row, self.COL_REQTYPE) or REQ_GRANT).strip()
                        self.refresh_notifications()
                        it.setText("제거완료" if req == REQ_RELEASE else "추가완료")
                        ok_cnt += 1
                    else:
                        it.setText(f"처리실패: {msg}")
                        fail_cnt += 1

            disp_no = (row + 1) if (0 <= row < self.table.rowCount()) else "?"
            if ok:
                lines.append(f"#{disp_no} BUS 완료 처리됨")
            else:
                reason = msg if msg else "실패"
                lines.append(f"#{disp_no} BUS 완료 처리 실패 ({reason})")

        if lines:
            self._log("\n".join(lines))
        self.status_label.setText(f"완료 처리 결과: 성공 {ok_cnt}건 / 실패 {fail_cnt}건")

        if self._bus_mode:
            self._bus_done += len(results or [])
            self.status_label.setText(f"완료 처리 중 {self._bus_done}/{self._bus_total}")
            self.refresh_notifications()
            self._start_next_bus_item()
            return

        if is_auto_wait:
            matched = any(res.get('row') == pending_row for res in (results or []))
            self._waiting_for_bus = False if matched else self._waiting_for_bus
            self._pending_after_add_row = None if matched else self._pending_after_add_row

            if self.run_queue:
                self._start_next_job()
            else:
                self._set_running_ui(False)
                self.status_label.setText("완료")
                self.refresh_notifications()

    def _on_session_ready(self, ok: bool, msg: str):
        if ok:
            self.status_label.setText("세션 준비됨")
        else:
            self.status_label.setText(f"세션 미준비: {msg}")

        self._set_running_ui(False)
        cbs = getattr(self, "_pending_bus_callbacks", [])
        self._pending_bus_callbacks = []

        for cb in cbs:
            try:
                cb(bool(ok))
            except Exception as e:
                self._log(f"[세션 준비 콜백 오류] {e}")

    def _on_session_busy(self, b: bool):
        self._set_running_ui(b)

    def _on_session_downloaded(self, path: str, err: str):
        self._set_running_ui(False)
        self.btn_request.setEnabled(True)
        self.btn_settings.setEnabled(True)
        
        if err:
            if err.strip() == "사용자 취소":
                self.status_label.setText("요청 확인 취소됨")
                return
            QMessageBox.critical(self, "오류", f"로드 실패: {err}")
            return
        
        self.status_label.setText("로드 완료")
        self.load_excel(path, append=False, silent=False)
        
        try:
            end_combined = os.path.join(DL_DIR, "종료권한리스트_합본.xls")
            normal_combined = os.path.join(DL_DIR, "권한리스트_합본.xls")
            if os.path.basename(path) == "권한리스트_합본.xls" and os.path.exists(end_combined):
                self.load_excel(end_combined, append=True, silent=True)
            elif os.path.basename(path) == "종료권한리스트_합본.xls" and os.path.exists(normal_combined):
                self.load_excel(normal_combined, append=True, silent=True)
        except Exception:
            pass

        prog_grant = prog_rel = end_grant = end_rel = 0
        for r in range(self.table.rowCount()):
            kind = (self._get(r, self.COL_KIND) or "").strip()
            req  = (self._get(r, self.COL_REQTYPE) or REQ_GRANT).strip()
            if   kind == "진행" and req == REQ_GRANT:   prog_grant += 1
            elif kind == "진행" and req == REQ_RELEASE: prog_rel   += 1
            elif kind == "종료" and req == REQ_GRANT:   end_grant  += 1
            elif kind == "종료" and req == REQ_RELEASE: end_rel    += 1

        parts = []
        if prog_grant: parts.append(f"진행-부여 {prog_grant}건")
        if prog_rel:   parts.append(f"진행-해제 {prog_rel}건")
        if end_grant:  parts.append(f"종료-부여 {end_grant}건")
        if end_rel:    parts.append(f"종료-해제 {end_rel}건")
        self._log("로드 완료: " + (" , ".join(parts) if parts else "항목 없음"))

    def _load_creds(self):
        self.creds = {"id":"", "pw":""}
        try:
            if os.path.exists(CONF_FILE):
                with open(CONF_FILE, "r", encoding="utf-8") as f:
                    self.creds = json.load(f)
                    self.creds.setdefault("debug", False)
        except:
            pass
        self.remove_fail_tolerance = int(self.creds.get("remove_fail_tol", 5))

    def _save_creds(self, id_, pw_, debug_):
        with open(CONF_FILE, "w", encoding="utf-8") as f:
            json.dump({"id": id_, "pw": pw_, "debug": bool(debug_), "theme": getattr(self, "current_theme", "light")},
                      f, ensure_ascii=False, indent=2)

    def closeEvent(self, e):
        try:
            if hasattr(self, "session") and self.session:
                self.trigger_session_cancel.emit()
                end = time.time() + 2.0
                while time.time() < end and self.session.is_busy():
                    QApplication.processEvents()
                    time.sleep(0.02)

            if hasattr(self, "session") and self.session:
                self.trigger_session_stop.emit()
                end2 = time.time() + 1.0
                while time.time() < end2:
                    QApplication.processEvents()
                    time.sleep(0.02)

            if hasattr(self, "session_thread") and self.session_thread:
                self.session_thread.quit()
                self.session_thread.wait(1500)

            if hasattr(self, "watch_session"):
                self.trigger_watcher_stop.emit()
                
            if hasattr(self, "watch_thread") and self.watch_thread:
                self.watch_thread.quit()
                self.watch_thread.wait(1500)

        except Exception:
            pass

        try:
            if hasattr(self, "session") and getattr(self.session, "driver", None):
                try:
                    self.session.driver.quit()
                except Exception:
                    pass

            if hasattr(self, "watch_session"):
                mgr = getattr(self.watch_session, "_mgr", None)
                drv = getattr(mgr, "driver", None) if mgr else None
                if drv:
                    try:
                        drv.quit()
                    except Exception:
                        pass
        except Exception:
            pass
        
        super().closeEvent(e)

    def _open_settings(self):
        current_debug = bool(self.creds.get("debug", False))
        remembered = bool(self.creds.get("id"))
        current_tol = int(self.creds.get("remove_fail_tol", 5))

        dlg = SettingsDialog(self, self.creds, remembered=remembered, debug_on=current_debug, fail_tol_default=current_tol)

        if dlg.exec_() == QDialog.Accepted:
            uid, pw, do_save, debug_on, fail_tol, notify_mins = dlg.result()
            tol = dlg.get_fail_tol()

            self.creds = {
                "id": uid if do_save else "",
                "pw": pw if do_save else "",
                "debug": bool(debug_on),
                "remove_fail_tol": int(tol),
                "notify_refresh_min": int(notify_mins),
            }

            if uid and pw:
                self.watch_session.set_creds(uid, pw)
                self.trigger_watcher_start.emit()
            else:
                self.trigger_watcher_stop.emit()

            data = {
                "theme": getattr(self, "current_theme", "light"),
                "debug": bool(debug_on),
                "remove_fail_tol": int(tol),
                "notify_refresh_min": int(notify_mins),
            }
            if do_save:
                data.update({"id": uid, "pw": pw})

            try:
                with open(CONF_FILE, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
            except Exception:
                pass

            self._notify_refresh_mins = int(notify_mins)
            self.notify_timer.setInterval(max(5, self._notify_refresh_mins) * 60 * 1000)
            self.debug_enabled = bool(debug_on)
            self.session.set_debug(self.debug_enabled, DEBUG_DIR)
            self.remove_fail_tolerance = int(tol)

    def _request_and_import(self):
        if not self.creds.get("id") or not self.creds.get("pw"):
            self._open_settings()
            if not self.creds.get("id") or not self.creds.get("pw"):
                QMessageBox.warning(self, "알림", "아이디/비밀번호가 필요합니다.")
                return

        self.session.set_creds(self.creds["id"], self.creds["pw"])

        if not self.session.is_ready():
            self._set_running_ui(True)
            self.status_label.setText("세션 준비 중")
            self.trigger_session_start.emit()

            end = time.time() + 60
            while time.time() < end and self.isVisible():
                QApplication.processEvents()
                if self.session.is_ready():
                    self.status_label.setText("세션 준비됨")
                    break
                time.sleep(0.05)

            if not self.session.is_ready():
                self._set_running_ui(False)
                QMessageBox.critical(self, "오류", "세션 초기화 실패")
                return

        self.btn_request.setEnabled(False)
        self.btn_settings.setEnabled(False)
        self._set_running_ui(True)
        self.status_label.setText("요청 확인 중")
        self.trigger_session_download.emit()

    def _clear_log(self):
        self.log.clear()

    def _stop_all(self):
        self.run_queue = []
        self.stop_requested = True
        
        try:
            if hasattr(self, "session"):
                self.trigger_session_cancel.emit()
        except Exception:
            pass
        if self.proc.state() == QProcess.Running:
            try:
                self.proc.terminate()
                self.proc.waitForFinished(1500)
                if self.proc.state() == QProcess.Running:
                    self.proc.kill()
                    self.proc.waitForFinished(1500)
            except Exception:
                pass

        for r in range(self.table.rowCount()):
            it = self.table.item(r, self.COL_STATUS)
            if it and (it.text().strip() == "실행중" or it.text().startswith("완료 처리중")):
                it.setText("중지됨")

        self._ignore_bus_results = True
        self._waiting_for_bus = False
        self._pending_after_add_row = None
        self._set_running_ui(False)
        self.status_label.setText("중지됨")
        self.btn_stop.setEnabled(False)

    def _set_running_ui(self, running: bool):
        self.prg.setVisible(running)
        self.status_label.setText("" if not running else f"처리중 {self.done_jobs}/{self.total_jobs}")

        to_disable = [
            getattr(self, "btn_manual", None),
            getattr(self, "btn_file", None),
            getattr(self, "btn_run_execute", None),
            getattr(self, "btn_run_complete", None),
            getattr(self, "btn_clear_log", None),
            getattr(self, "btn_request", None),
            getattr(self, "btn_settings", None),
            getattr(self, "chk_dry", None),
            getattr(self, "chk_auto_complete", None),
            getattr(self, "btn_newcheck", None),
        ]
        for w in to_disable:
            if w:
                w.setEnabled(not running)

        if hasattr(self, "btn_stop") and self.btn_stop:
            self.btn_stop.setEnabled(running)

        if hasattr(self, "table") and self.table:
            self.table.setContextMenuPolicy(Qt.PreventContextMenu if running else Qt.CustomContextMenu)

    def _wrap_cmd_utf8(self, cmd: str) -> str:
        pre = "$OutputEncoding=[System.Text.Encoding]::UTF8; [Console]::OutputEncoding=[System.Text.Encoding]::UTF8;"
        if self.ps_kind == "powershell":
            pre = pre + " chcp 65001 > $null;"
        return pre + " " + cmd

    def _pick_powershell(self):
        p = shutil.which("pwsh")
        if p: return p, "pwsh"
        p = shutil.which("powershell")
        return (p if p else "powershell"), "powershell"

    def _start_next_job(self):
        if self.stop_requested or not self.run_queue:
            self._set_running_ui(False)
            return
        
        (seq, row, cmd, pretty, mode, reqtype, user, proj, lv2, lv3, path) = self.run_queue.pop(0)
        self.run_queue_cmd = cmd
        self.run_queue_pretty = pretty
        self.current_mode = mode
        self.current_row = row
        self.current_seq = seq
        kind = (self._get(row, self.COL_KIND) or "").strip()    
        self._current_reqtype = reqtype
        self._current_target = {'row': row, 'kind': kind, 'user': user, 'proj': proj, 'lv2': lv2, 'lv3': lv3, 'path': path, 'req': reqtype}
        is_closed = (not (lv2 or "").strip()) and (not (lv3 or "").strip())
        if is_closed:
            self._log(f"#{seq} {mode.upper()} {user} / {proj} 시작")
        else:
            self._log(f"#{seq} {mode.upper()} {user} / {proj} / {lv2} / {lv3} 시작")

        itm = self.table.item(row, self.COL_STATUS)
        if itm:
            itm.setText("실행중")
        self.table.blockSignals(False)

        self._cur_cmd = cmd
        self._cur_pretty = pretty

        wrapped = self._wrap_cmd_utf8(cmd)
        args = ["-NoLogo","-NoProfile","-ExecutionPolicy","Bypass","-Command", wrapped]
        self._set_running_ui(True)
        self._buf_out[self.current_seq] = []
        self.proc.start(self.ps_path, args)
        self.status_label.setText(f"처리중 {self.done_jobs}/{self.total_jobs}")

    def _filter_log_lines(self, mode: str, raw: str):
        lines = [ (line or "").strip() for line in raw.splitlines() if (line or "").strip() ]

        if mode == "remove":
            keep = []

            for s in lines:
                if ("ERROR" in s or "Error" in s or 
                    "Access is denied" in s or "Access denied" in s or "denied" in s.lower()):
                    keep.append(s)

            summary_re = re.compile(r"Successfully processed\s+\d+\s+files;?\s+Failed processing\s+\d+\s+files", re.IGNORECASE)
            summaries = [s for s in lines if summary_re.search(s)]
            if summaries:
                keep.append(summaries[-1])

            return keep

        keep = []
        for s in lines:
            if any(k in s for k in (
                "ERROR", "Error", "Access is denied", "denied", "Access denied",
                "processed file:", "Successfully processed", "Failed processing"
            )) or re.search(r'\b(Success|Fail(ed)?|Denied|removed|grant(ed)?)\b', s, re.I):
                keep.append(s)
        return keep

    def _ps_ready_out(self):
        raw = bytes(self.proc.readAllStandardOutput()).decode("utf-8", errors="ignore")
        if not raw:
            return
        lines = self._filter_log_lines(self.current_mode, raw)
        if lines:
            tagged = [f"#{self.current_seq} {line}" for line in lines]
            self._buf_out.setdefault(self.current_seq, []).extend(tagged)

    def _ps_ready_err(self):
        raw = bytes(self.proc.readAllStandardError()).decode("utf-8", errors="ignore")
        if not raw:
            return
        lines = self._filter_log_lines(self.current_mode, raw)
        if not lines:
            lines = [x.strip() for x in raw.splitlines() if x.strip()]
        if lines:
            tagged = [f"#{self.current_seq} {line}" for line in lines]
            self._buf_out.setdefault(self.current_seq, []).extend(tagged)

    def _ps_finished(self, code, status):
        try:
            buf = self._buf_out.pop(self.current_seq, [])
        except Exception:
            buf = []

        is_remove = (self.current_mode == "remove")

        fail_cnt = 0
        succ_cnt = 0
        if is_remove:
            for line in buf:
                m_fail = re.search(r'Failed processing\s+(\d+)', line, re.I)
                if m_fail:
                    fail_cnt += int(m_fail.group(1))
                m_succ = re.search(r'Successfully processed\s+(\d+)', line, re.I)
                if m_succ:
                    succ_cnt = max(succ_cnt, int(m_succ.group(1)))

            if fail_cnt == 0:
                extra_fails = [l for l in buf if any(k in l for k in (
                    "Failed processing", "ERROR", "Error", "Access is denied", "Access denied", "denied"
                ))]
                fail_cnt = len(extra_fails)

            tol = int(getattr(self, "remove_fail_tolerance", 5))
            self._log(f"#{self.current_seq} 제거 요약: 성공 {succ_cnt} / 실패 {fail_cnt} (허용 {tol})")

            effective_code = 0 if (fail_cnt <= tol) else 1

        else:
            if buf:
                self._log("\n".join(buf))
            effective_code = 0 if code == 0 else 1

        missing_group = False
        buf_text = "\n".join([b.split("#", 1)[-1].strip() if "#" in b else b for b in (buf or [])])

        if self.current_mode == "add" and effective_code != 0:
            if ("개체를 찾을 수 없습니다" in buf_text) or ("Cannot find an object with identity" in buf_text):
                missing_group = True

        if missing_group and not getattr(self, "_retrying_after_group_create", False):
            do_create = False

            if getattr(self, "_auto_create_group_decided", None) is True:
                do_create = True
            else:
                mb = QMessageBox(self)
                mb.setWindowTitle("보안그룹 생성")
                mb.setIcon(QMessageBox.Question)
                gname = format_group_name(self._current_target.get('proj',''), self._current_target.get('lv2',''))
                mb.setText(f"보안그룹 '{gname}' 이(가) 없습니다.\n생성 후 계속 진행할까요?")
                btn_continue = mb.addButton("계속", QMessageBox.AcceptRole)
                btn_cancel   = mb.addButton("취소", QMessageBox.RejectRole)
                cb = QCheckBox("이번 세션 동안 자동 생성(다시 묻지 않음)", mb)
                mb.setCheckBox(cb)
                mb.exec_()
                do_create = (mb.clickedButton() == btn_continue)
                if do_create and cb.isChecked():
                    self._auto_create_group_decided = True

            if do_create:
                ok_create, msg_create = self._create_group_and_base_acl(
                    proj=self._current_target.get('proj',''),
                    lv2=self._current_target.get('lv2',''),
                    lv3=self._current_target.get('lv3','')
                )
                if ok_create:
                    self._retrying_after_group_create = True

                    self.run_queue.insert(0, (
                        self.current_seq, self.current_row,
                        getattr(self, "_cur_cmd", ""), getattr(self, "_cur_pretty", ""),
                        self.current_mode, self._current_reqtype,
                        self._current_target.get('user'), self._current_target.get('proj'),
                        self._current_target.get('lv2'), self._current_target.get('lv3'),
                        self._current_target.get('path')
                    ))

                    self._start_next_job()
                    return
                else:
                    self._log(f"[보안그룹 생성 실패] {msg_create}")

        if self.current_mode != "remove":
            req_label = "권한 부여" if self._current_reqtype == REQ_GRANT else "권한 제거"
            if effective_code == 0:
                self._log(f"#{self.current_seq} {req_label} 완료")
            else:
                reason = "실패"
                try:
                    reason = (buf[-1].split("#", 1)[-1].strip() if buf else "실패")
                except Exception:
                    pass
                self._log(f"#{self.current_seq} {req_label} 실패 ({reason})")

        self.done_jobs += 1

        if 0 <= self.current_row < self.table.rowCount():
            if effective_code == 0:
                ok = "추가완료" if self.current_mode == "add" else "제거완료"
                self.table.item(self.current_row, self.COL_STATUS).setText(ok)

                if (self.auto_complete_after_add
                    and not self.chk_dry.isChecked()
                    and self.session.is_ready()
                    and not self.stop_requested):
                    try:
                        tgt = dict(self._current_target)
                        tgt["req"] = REQ_GRANT if self.current_mode == "add" else REQ_RELEASE
                        self._waiting_for_bus = True
                        self._pending_after_add_row = self.current_row
                        self.table.item(self.current_row, self.COL_STATUS).setText("완료 처리중")
                        self.trigger_session_process.emit([tgt])
                    except Exception as e:
                        self._log(f"[완료처리 요청 실패] row={self.current_row+1 if self.current_row>=0 else '?'} / {e}")
                        self._waiting_for_bus = False
                        self._pending_after_add_row = None
            else:
                self.table.item(self.current_row, self.COL_STATUS).setText("실패")

        self.status_label.setText(f"처리중 {self.done_jobs}/{self.total_jobs}")
        self._retrying_after_group_create = False
        if not self._waiting_for_bus:
            if self.run_queue:
                self._start_next_job()
            else:
                self.status_label.setText("완료" if not getattr(self, "stop_requested", False) else "중지됨")
                self._set_running_ui(False)
                self.stop_requested = False

    def open_manual_dialog(self):
        try:
            dlg = ManualEntryDialog(self)
            if dlg.exec_() == QDialog.Accepted and dlg.result_row:
                reqtype, user, proj, lv2, lv3, dept, role, kind = dlg.result_row
                self.add_table_row(reqtype, user, proj, lv2, lv3, dept, role, kind)   
        except Exception as e:
            self._log(f"[수동입력 오픈 예외] {e}")
            QMessageBox.critical(self, "오류", f"수동 입력 다이얼로그 실행 오류:\n{e}")

    def add_table_row(self, reqtype: str, user: str, proj: str, lv2: str, lv3: str, dept: str, role: str, kind: str = "진행"):
        self.table.blockSignals(True)
        try:
            r = self.table.rowCount()
            self.table.insertRow(r)

            chk = QCheckBox()
            chk.setChecked(True)
            
            wrapper = QWidget()
            layout = QHBoxLayout(wrapper)
            layout.addWidget(chk)
            layout.setAlignment(Qt.AlignCenter)
            layout.setContentsMargins(0,0,0,0)
            self.table.setCellWidget(r, self.COL_SELECT, wrapper)

            vals = [kind, reqtype or REQ_GRANT, user, "", proj, lv2, lv3, dept or "", role or "", "대기"]
            
            for j, val in enumerate(vals, start=1):
                col = j
                it = QTableWidgetItem(str(val))
                it.setTextAlignment(Qt.AlignCenter)

                if col in self.EDITABLE_COLS:
                    it.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)
                else:
                    it.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                self.table.setItem(r, col, it)
        finally:
            self.table.blockSignals(False)

        if (kind or "").strip() == "종료":
            self._log(f"수동 입력 추가: 구분={kind}, 요청={reqtype}, 사번={user}, 프로젝트={proj}")
        else:
            self._log(f"수동 입력 추가: 구분={kind}, 요청={reqtype}, 사번={user}, 프로젝트={proj}, L2={lv2}, L3={lv3}, ROLE={'없음' if not role else role}")

    def _log(self, message: str, seq: int = None, dry: bool = False):
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        dry_str = " Dry" if dry else ""
        tag = f"{dry_str} #{seq}" if seq is not None else ""

        norm = (message or "").replace("\r", "\n").replace("\t", "    ").rstrip()
        block = f"[{ts}{tag}]\n{norm}\n"

        cursor = self.log.textCursor()
        cursor.movePosition(QTextCursor.End)
        self.log.setTextCursor(cursor)

        if self.log.toPlainText():
            self.log.insertPlainText("\n")

        self.log.insertPlainText(block)
        self.log.ensureCursorVisible()

        fn = os.path.join(ACCESS_LOG_DIR, f"access_{datetime.date.today().strftime('%Y%m%d')}.log")
        with open(fn, "a", encoding="utf-8") as f:
            f.write(block)

    def choose_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "승인 엑셀 선택", "", "Excel Files (*.xlsx *.xls)")
        if not path:
            return
        self.load_excel(path)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            for url in event.mimeData().urls():
                lf = url.toLocalFile().lower()
                if lf.endswith(".xlsx") or lf.endswith(".xls"):
                    event.acceptProposedAction(); return
        event.ignore()

    def dropEvent(self, event):
        urls = event.mimeData().urls()
        if not urls: return
        path = urls[0].toLocalFile()
        if path.lower().endswith((".xlsx",".xls")):
            self.load_excel(path)

    def load_excel(self, file_path: str, append: bool = False, silent: bool = False):
        try:
            self.table.blockSignals(True)
            #if not (append and silent):
                #self.file_label.setText(file_path)
            ext = os.path.splitext(file_path)[1].lower()

            start_offset = self.table.rowCount() if append else 0
            
            rows = []

            if ext == ".xlsx":
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                header_raw = [str(c.value) if c.value is not None else "" for c in ws[1]]
                colmap = auto_map_columns(header_raw)
                
                is_end = any(("열람" in (str(h) or "")) for h in header_raw)
                kind = "종료" if is_end else "진행"
                
                required = {"user", "proj"} if is_end else {"user", "proj", "level2", "level3"}
                missing = [k for k in required if k not in colmap]
                if missing:
                    self._log(f"엑셀 로드 실패: 필요한 컬럼 없음 -> {missing} / 헤더: {header_raw}")
                    return

                def gv(row, key, default=""):
                    i = colmap.get(key)
                    if i is None: 
                        return default
                    v = row[i] if i < len(row) else None
                    return (str(v).strip() if v is not None else default)
                
                for r in ws.iter_rows(min_row=2, values_only=True):
                    if not any(r):
                        continue
                    user_id = gv(r, "user")
                    name    = gv(r, "name")
                    proj    = gv(r, "proj")
                    lv2     = "" if is_end else gv(r, "level2")
                    lv3     = "" if is_end else gv(r, "level3")
                    role    = gv(r, "role") if not is_end else ""
                    dept    = gv(r, "dept")
                    
                    if not user_id or not proj:
                        continue
                    if not is_end and (not lv2 or not lv3):
                        continue
                    
                    reqtype = REQ_RELEASE if _is_release_row_by_values(list(r), header_raw) else REQ_GRANT
                    path = build_path_l3(proj, lv2, lv3)
                    rows.append((kind, reqtype, user_id, name, proj, lv2, lv3, dept, role, "대기"))

            elif ext == ".xls":
                header_raw, data_rows = _parse_html_best_table(file_path)
                if not header_raw:
                    self._log("엑셀 로드 실패: .xls(HTML) 테이블을 찾지 못했습니다.")
                    return
                
                colmap = auto_map_columns(header_raw)

                is_end = any(("열람" in (str(h) or "")) for h in header_raw)
                kind = "종료" if is_end else "진행"
                
                required = {"user", "proj"} if is_end else {"user", "proj", "level2", "level3"}
                missing = [k for k in required if k not in colmap]
                if missing:
                    self._log(f"엑셀 로드 실패: 필요한 컬럼 없음 -> {missing} / 헤더: {header_raw}")
                    return

                def gv(row, key, default=""):
                    i = colmap.get(key)
                    if i is None: 
                        return default
                    v = row[i] if i < len(row) else None
                    return (str(v).strip() if v is not None else default)

                for r in data_rows:
                    user_id = gv(r, "user")
                    name    = gv(r, "name")
                    proj    = gv(r, "proj")
                    lv2     = "" if is_end else gv(r, "level2")
                    lv3     = "" if is_end else gv(r, "level3")
                    role    = gv(r, "role") if not is_end else ""
                    dept    = gv(r, "dept")
                    
                    if not user_id or not proj:
                        continue
                    if not is_end and (not lv2 or not lv3):
                        continue
                    
                    reqtype = REQ_RELEASE if _is_release_row_by_values(r, header_raw) else REQ_GRANT
                    path = build_path_l3(proj, lv2, lv3)
                    rows.append((kind, reqtype, user_id, name, proj, lv2, lv3, dept, role, "대기"))

            else:
                self._log(f"엑셀 로드 실패: 지원하지 않는 확장자 ({ext})")
                return

            if append:
                self.table.setRowCount(start_offset + len(rows))
                base = start_offset
            else:
                self.table.setRowCount(len(rows))
                base = 0

            for i, row in enumerate(rows):
                r = base + i
                
                chk = QCheckBox()
                chk.setChecked(True)
                
                wrapper = QWidget()
                layout = QHBoxLayout(wrapper)
                layout.addWidget(chk)
                layout.setAlignment(Qt.AlignCenter)
                layout.setContentsMargins(0, 0, 0, 0)
                self.table.setCellWidget(r, self.COL_SELECT, wrapper)

                for j, val in enumerate(rows[i]):
                    col = j + 1
                    it = QTableWidgetItem(str(val))
                    it.setTextAlignment(Qt.AlignCenter)
                    if col in self.EDITABLE_COLS:
                        it.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)
                    else:
                        it.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    self.table.setItem(r, col, it)

            hv = self.table.horizontalHeader()
            hv.setStretchLastSection(True)

        except Exception as e:
            self._log(f"엑셀 로드 실패: {e}")

        finally:
            self.table.blockSignals(False)

    def _on_cell_changed(self, row, col):
        if col == self.COL_STATUS:
            return

        if col in self.EDITABLE_COLS:
            st = self.table.item(row, self.COL_STATUS)
            if st:
                st.setText("검증필요")
        
    def _open_table_menu(self, pos):
        row = self.table.indexAt(pos).row()
        m = QMenu(self)
        act_del_row = m.addAction("행 삭제")
        act_del_sel = m.addAction("선택 행 삭제")
        act_del_all = m.addAction("전체 삭제")
        
        gpos = self.table.viewport().mapToGlobal(pos)
        act = m.exec_(gpos)
        
        if act == act_del_row:
            self._delete_row(row)
        elif act == act_del_sel:
            self._delete_checked_rows()
        elif act == act_del_all:
            self._delete_all_rows()     

    def _delete_all_rows(self):
        self.table.blockSignals(True)
        try:
            self.table.setRowCount(0)
        finally:
            self.table.blockSignals(False)

    def _delete_row(self, row: int):
        if row < 0 or row >= self.table.rowCount():
            return
        self.table.removeRow(row)

    def _get_checkbox(self, row: int) -> QCheckBox:
        w = self.table.cellWidget(row, self.COL_SELECT)
        if not w:
            return None
        return w.findChild(QCheckBox)

    def _is_row_checked(self, row: int) -> bool:
        cb = self._get_checkbox(row)
        return bool(cb and cb.isChecked())

    def _delete_checked_rows(self):
        for r in range(self.table.rowCount()-1, -1, -1):
            if self._is_row_checked(r):
                self.table.removeRow(r)

    def validate_row(self, row: int, mode: str = "add") -> (bool, str):
        kind = (self._get(row, self.COL_KIND) or "").strip()
        proj = self._get(row, self.COL_PROJ).strip()

        if kind == "종료":
            if not proj:
                return False, f"{row+1}행: 종료 과제 실행에 필요한 필드(proj) 누락"
            return True, ""

        role = self._get(row, self.COL_ROLE).strip()
        lv2  = self._get(row, self.COL_LV2).strip()
        lv3  = self._get(row, self.COL_LV3).strip()

        if mode == "remove":
            if not (proj and lv2 and lv3):
                return False, f"{row+1}행: 제거 실행에 필요한 필드(proj/lv2/lv3) 누락"
            return True, ""

        lv2_norm = normalize_lv2(lv2)
        is_new = is_new_template(proj)

        if not role:
            return True, ""

        if lv2_norm == "Study":
            if role not in STUDY_ROLES:
                return False, f"{row+1}행: Study에서는 허용되지 않는 STATROLE '{role}'"
            if is_new and role not in ROLE_MAP:
                return False, f"{row+1}행: 신버전 Study에서 STATROLE '{role}' 매핑 없음(ROLE_MAP 보강 필요)"
            if not is_new:
                needed = _legacy_needed_nums(role)
                stat_path = build_path_l3(proj, lv2, lv3)
                missing = _legacy_find_missing_dirs(stat_path, needed)
                if missing:
                    miss_str = ", ".join(map(str, missing))
                    return False, (
                        f"{row+1}행: 레거시 Study 폴더 부족 → 필요한 번호 폴더({miss_str}) 없음\n"
                        f"경로: {stat_path}\n조치: 폴더 생성/정정 후 다시 실행하세요."
                    )

        elif lv2_norm == "Isolated":
            if is_stat_idmc_lv3(lv3):
                if is_stat_idmc_new_policy(proj):
                    if role not in ISOLATED_STAT_IDMC_ROLE_MAP:
                        return False, f"{row+1}행: STAT_IDMC 폴더에 허용되지 않는 Role. 확인 필요. ('{role} = ROLE_MAP 매핑 없음')"
            elif is_new:
                if role not in ISOLATED_ROLES:
                    return False, f"{row+1}행: Isolated에서는 허용되지 않는 STATROLE '{role}'"
            else:
                if role != "Randomization Statistician":
                    return False, f"{row+1}행: 과거 Isolated 폴더는 'Randomization Statistician'만 가능 (현재 '{role}')"
                iso_path = build_path_l3(proj, lv2, lv3)
                if not os.path.isdir(iso_path):
                    return False, f"{row+1}행: Isolated 경로 없음 → {iso_path}\n조치: 경로 확인/생성 후 다시 실행하세요."
        else:
            return False, f"{row+1}행: Level2 '{lv2}'에서는 STATROLE 사용 불가"

        return True, ""

    def toggle_all_rows(self, checked: bool):
        for r in range(self.table.rowCount()):
            cb = self._get_checkbox(r)
            if cb:
                cb.setChecked(checked)

    def run_execute(self):
        self._ignore_bus_results = False
        self._waiting_for_bus = False
        self._pending_after_add_row = None
        self.stop_requested = False
        self.auto_complete_after_add = self.chk_auto_complete.isChecked()

        if self.table.rowCount() == 0:
            self._log("실행할 데이터가 없습니다."); return
        if self.session.is_busy():
            QMessageBox.warning(self, "알림", "다른 작업이 실행 중입니다. 잠시 후 다시 시도하세요."); return

        dry = self.chk_dry.isChecked()
        seq = 0
        failed_msgs = []
        queue = []

        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, self.COL_SELECT)
            cb = w.findChild(QCheckBox) if w else None
            if not (cb and cb.isChecked()):
                continue

            kind    = (self._get(r, self.COL_KIND) or "").strip()
            reqtype = (self._get(r, self.COL_REQTYPE) or REQ_GRANT).strip()
            mode    = "add" if reqtype == REQ_GRANT else "remove"

            ok, reason = (True, "")
            if mode == "add":
                ok, reason = self.validate_row(r, mode="add")
            else:
                ok, reason = self.validate_row(r, mode="remove")

            if not ok:
                self.table.item(r, self.COL_STATUS).setText("검증실패")
                short = reason.splitlines()[0]
                self.table.item(r, self.COL_STATUS).setText(f"검증실패: {short}")
                self._log(reason)
                failed_msgs.append(reason)
                continue

            user = self._get(r, self.COL_USER)
            proj = self._get(r, self.COL_PROJ)
            lv2  = self._get(r, self.COL_LV2)
            lv3  = self._get(r, self.COL_LV3)
            role = self._get(r, self.COL_ROLE)
            
            if kind == "종료":
                path = build_closed_path_from_proj(proj)
                cmd  = generate_add_script_closed(user, proj) if mode == "add" else generate_remove_script_closed(user, proj)
            else:
                path = build_path_l3(proj, lv2, lv3)
                cmd  = generate_add_script(user, proj, lv2, lv3, role) if mode == "add" else generate_remove_script(user, proj, lv2, lv3, role)

            body = pretty_cmd_lines(cmd) if cmd else "(생성된 명령 없음)"

            if dry:
                seq += 1
                self._log(body, seq=seq, dry=True)
                self.table.blockSignals(True)
                self.table.item(r, self.COL_STATUS).setText("DryRun")
                self.table.blockSignals(False)
                continue

            if not cmd:
                continue

            seq += 1
            queue.append((seq, r, cmd, body, mode, reqtype, user, proj, lv2, lv3, path))

        if failed_msgs:
            QMessageBox.warning(self, "검증 실패 요약", "아래 항목은 검증 실패로 실행하지 않았습니다.\n\n" + "\n\n".join(failed_msgs))

        if not queue:
            return

        self.total_jobs = len(queue)
        self.done_jobs = 0
        self._set_running_ui(True)
        self.run_queue = queue
        self._start_next_job()

    def _get(self, row: int, col: int) -> str:
        it = self.table.item(row, col)
        return it.text().strip() if it else ""

def _excepthook(etype, value, tb):
    try:
        import traceback
        msg = "".join(traceback.format_exception(etype, value, tb))
        print(msg)
    except:
        pass

    QMessageBox.critical(None, "치명적 오류", str(value))

import sys as _sys
_sys.excepthook = _excepthook

if __name__ == "__main__":
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    
    app = QApplication(sys.argv)
    app.setFont(QFont("Segoe UI", 9))
    window = AccessManager()
    def _on_quit():
        try:
            window.close()
        except Exception:
            pass
    window.show()
    sys.exit(app.exec_())
